from rich.console import Console
from rich.table import Table
from rich.live import Live
from pathlib import Path
import pdfkit
import pandas as pd
import os
from dominate.tags import *
from dominate.util import raw
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
import json
import time


class FormatConfig:
    def __init__(self, config_file):
        with open(config_file) as jf:
            self.formatting = json.load(jf)

    def load_body_format(self):
        return self.formatting["body"]

    def load_table_format(self):
        return self.formatting["table"]

    def load_rpm_info_format(self):
        return self.formatting['highlight']['rpm-info']

    def load_support_flag_format(self):
        return self.formatting['highlight']['support-flag']

    def load_running_format(self):
        return self.formatting['highlight']['running']


class HTMLTableFormatting:
    def __init__(self, formatting_config_file='config/color.default.json'):
        self.formatting = FormatConfig(formatting_config_file)

    def get_style(self):
        table_formatting = self.formatting.load_table_format()
        styles = [
            dict(selector='table', props=[('border', table_formatting['border'])]),
            dict(selector='th', props=[('border', table_formatting['th']['border']),
                                    ('font-size', table_formatting['th']['font-size']),
                                   ('font-family', table_formatting['th']['font-family']),
                                   ('background-color', table_formatting['th']['background-color'])]),
            dict(selector='td', props=[('border', table_formatting['th']['border']),
                                    ('font-size', table_formatting['td']['font-size']),
                                   ('font-family', table_formatting['td']['font-family'])])
        ]

        return styles

   

class RPMsExporter:
    def __init__(self, logger, formatting_config_file='config/color.default.json'):
        self.logger = logger
        self.formatting = FormatConfig(formatting_config_file)
        # self.console = Console()
        self.table = Table(show_header=True, header_style='bold green', show_lines=True)
        self.table.add_column('Name', width=32)
        self.table.add_column('Path', width=64)
        self.table.add_column('Vendor', width=12)
        self.table.add_column('Signature', width=28)
        self.table.add_column('Distribution', width=12)
        self.table.add_column('Driver Flag: supported')

    def to_terminal(self, rpm_table):
        with Live(self.table):
            for _, row in rpm_table.iterrows():
                name = row['Name']
                path = row['Path']
                vendor = row['Vendor']
                signature = row['Signature']
                distribution = row['Distribution']
                driver_supported = row['Driver Flag: supported']

                ds_formatting = ''
                for driver in driver_supported.split('\n'):
                    values = driver.split(': ')

                    if len(values) < 2:
                        ds_formatting += '[red]' + driver + '[/red]' + '\n'
                    elif values[1] == 'yes':
                        ds_formatting += '[green]' + driver + '[/green]' + '\n'
                    elif values[1] == 'external':
                        ds_formatting += '[blue]' + driver + '[/blue]' + '\n'
                    elif values[1] == 'no' or values[1] == 'Missing':
                        ds_formatting += '[red]' + driver + '[/red]' + '\n'

                self.table.add_row(name, path, vendor, signature, distribution, ds_formatting)


    def supported_formatting(self, value):
        formatting = self.formatting.load_support_flag_format()
        if ': no' in value or ': Missing' in value:
            return 'background-color: ' + formatting['Missing']['background-color']
        elif ': yes' in value:
            return 'background-color: ' + formatting['yes']['background-color']
        elif ': external' in value:
            return 'background-color: ' + formatting['external']['background-color']
        else:
            return 'background-color: white'
    

    def to_html(self, rpm_table, file):
        tableFormatter = HTMLTableFormatting()
        styles = tableFormatter.get_style()

        rpm_table['Driver Flag: supported'] = rpm_table['Driver Flag: supported'].str.replace('\n', '</br>')

        s = rpm_table.style.applymap(self.supported_formatting,
                          subset=pd.IndexSlice[:, ['Driver Flag: supported']],
                          ).hide_index()

        s = s.set_table_styles(styles)

        with open(file, 'w') as f:
            f.write(s.render())

    def to_excel(self, rpm_table, file):
        writer = pd.ExcelWriter(file, engine='xlsxwriter')
        rpm_table.to_excel(writer, index=False, sheet_name='Solid Driver Checks')
        workbook = writer.book

        supported_formatting = self.formatting.load_support_flag_format()
        worksheet = writer.sheets['Solid Driver Checks']
        na_format = workbook.add_format({'bg_color': '\'' + supported_formatting['Missing']['background-color'] + '\''})
        yes_format = workbook.add_format({'bg_color': '\'' + supported_formatting['yes']['background-color'] + '\''})
        external_format = workbook.add_format({'bg_color': '\'' + supported_formatting['external']['background-color'] + '\''})

        area = 'E2:E' + str(len(df.index) + 1)
        worksheet.conditional_format(area, {'type': 'text',
                                        'criteria': 'containing',
                                        'value': ': Missing',
                                        'format': na_format})
        worksheet.conditional_format(area, {'type': 'text',
                                        'criteria': 'containing',
                                        'value': ': external',
                                        'format': external_format})
        worksheet.conditional_format(area, {'type': 'text',
                                        'criteria': 'containing',
                                        'value': ': yes',
                                        'format': yes_format})

        writer.save()

    def to_pdf(self, rpm_table, file):
        self.to_html(rpm_table, '.tmp.rpms.html')
        pdfkit.from_file('.tmp.rpms.html', file)
        os.remove('.tmp.rpms.html')

    def to_all(self, rpm_table, directory):
        self.to_html(rpm_table, os.path.join(directory, 'check_result.html'))
        self.to_pdf(rpm_table, os.path.join(directory, 'check_result.pdf'))
        self.to_excel(rpm_table, os.path.join(directory, 'check_result.xlsx'))


class DriversExporter:
    def __init__(self, logger, formatting_config_file='config/color.default.json'):
        self.logger = logger
        self.formatting = FormatConfig(formatting_config_file)
        

    def supported_color(self, value):
        if value is 'yes':
            return '[green]' + value + '[/green]'
        elif value is 'external':
            return '[blue]' + value + '[/blue]'
        else:
            return '[red]' + value + '[/red]'

    def running_color(self, value):
        if value is 'True':
            return '[green]True[/green]'
        else:
            return '[gray]False[gray]'
        

    def rpm_info_color(self, value):
        if 'is not owned by any package' in value:
            return '[red]' + value + '[/red]'
        else:
            return value
        

    def to_terminal(self, driver_tables):
        for server in driver_tables:
            table = Table(show_header=True, header_style='bold green', show_lines=True)
            table.add_column('Name', width=32)
            table.add_column('Path', width=64)
            table.add_column('SUSE Release', width=32)
            table.add_column('Flag: supported', width=28)
            table.add_column('Running', width=12)
            table.add_column('RPM Information')

            data = driver_tables[server]
            supported_formatting = self.formatting.load_support_flag_format()
            running_formatting = self.formatting.load_running_format()
            rpm_info_formatting = self.formatting.load_rpm_info_format()

            with Live(table):
                for _, row in data.iterrows():
                    supported = self.supported_color(row['Flag: supported'])
                    running = self.running_color(row['Running'])
                    rpm_info = self.rpm_info_color(row['RPM Information'])

                    table.add_row(row['Name'], row['Path'], row['SUSE Release'], supported, running, rpm_info)


    def supported_html_format_handler(self, value):
        value = value.lstrip().rstrip()
        supported = self.formatting.load_support_flag_format()
        if value is 'yes':
            return  'background-color:%s' % supported['yes']['background-color']
        elif value is 'external':
            return 'background-color:%s' % supported['external']['background-color']
        elif value is 'Missing' or value is 'no':
            return 'background-color:%s' % supported['Missing']['background-color']

        return ''

    def running_html_format_handler(self, value):
        running_format = self.formatting.load_running_format()
        if value is 'True':
            return 'background-color:%s' % running_format['true']['background-color']

        return 'background-color:%s' % running_format['false']['background-color']

    def rpm_info_html_format_handler(self, value):
        rpm_format = self.formatting.load_rpm_info_format()
        if 'is not owned by any package' in value:
            return 'background-color:%s' % rpm_format['no-rpm']['background-color']

        return ''

    def to_html(self, driver_tables, file):
        table_formatting = self.formatting.load_table_format()
        html_table_formatter = HTMLTableFormatting()
        styles = html_table_formatter.get_style()
        body_format = self.formatting.load_body_format()
        context = html()
        with context:
            font_family = body_format["font-family"]
            body_style = 'font-family: %s;' % font_family
            with body(style=body_style):
                for key in driver_tables:
                    h1('Solid Driver Checking Result: ' + key)

                    s = driver_tables[key].style.applymap(self.supported_html_format_handler,
                                subset=pd.IndexSlice[:, ['Flag: supported']])
                    s = s.applymap(self.running_html_format_handler,
                               subset=pd.IndexSlice[:, ['Running']])
                    s = s.applymap(self.rpm_info_html_format_handler,
                               subset=pd.IndexSlice[:, ['RPM Information']]).hide_index()
                    s = s.set_table_styles(styles)

                    div(raw(s.render()))

        with open(file, 'w') as f:
            f.write(context.render())


    def to_excel(self, driver_tables, file):
        for server in driver_tables:
            writer = pd.ExcelWriter(file, engine='openpyxl')
            if os.path.exists(file):
                writer = pd.ExcelWriter(file, engine='openpyxl', mode='a')

            driver_tables[server].to_excel(writer, index=False, sheet_name=server)
            writer.save()
            writer.close()

            workbook = load_workbook(filename=file)
            worksheet = workbook[server]

            records = str(len(driver_tables[server].index) + 1)
            support_flag_area = 'C2:C' + records
            running_area = 'E2:E' + records
            rpm_info_area = 'F2:F' + records

            support_flag_format = self.formatting.load_support_flag_format()
            na_text = Font(color=support_flag_format['Missing']['font-color'])
            na_fill = PatternFill(bgColor=support_flag_format['Missing']['background-color'])
            na = DifferentialStyle(font=na_text, fill=na_fill)
            yes_text = Font(color=support_flag_format['yes']['font-color'])
            yes_fill = PatternFill(bgColor=support_flag_format['yes']['background-color'])
            yes = DifferentialStyle(font=yes_text, fill=yes_fill)
            external_text = Font(color=support_flag_format['external']['font-color'])
            external_fill = PatternFill(bgColor=support_flag_format['external']['background-color'])
            external = DifferentialStyle(font=external_text, fill=external_fill)
            running_format = self.formatting.load_running_format()
            false_text = Font(color=running_format['false']['font-color'])
            false_fill = PatternFill(bgColor=running_format['false']['background-color'])
            false_format = DifferentialStyle(font=false_text, fill=false_fill)
            true_text = Font(color=running_format['true']['font-color'])
            true_fill = PatternFill(bgColor=running_format['true']['background-color'])
            true_format = DifferentialStyle(font=true_text, fill=true_fill)
            rpm_info_format = self.formatting.load_rpm_info_format()
            no_rpm_text = Font(color=rpm_info_format['no-rpm']['font-color'])
            no_rpm_fill = PatternFill(bgColor=rpm_info_format['no-rpm']['background-color'])
            no_rpm = DifferentialStyle(font=no_rpm_text, fill=no_rpm_fill)
            support_flag_na_rule = Rule(type='containsText', operator='containsText', text='Missing', dxf=na)
            support_flag_external_rule = Rule(type='containsText', operator='containsText', text='external', dxf=external)
            support_flag_yes_rule = Rule(type='containsText', operator='containsText', text='yes', dxf=yes)
            running_yes_rule = Rule(type='containsText', operator='containsText', text='True', dxf=true_format)
            running_no_rule = Rule(type='containsText', operator='containsText', text='False', dxf=false_format)
            no_rpm_rule = Rule(type='containsText', operator='containsText', text='is not owned by any package', dxf=no_rpm)

            worksheet.conditional_formatting.add(support_flag_area, support_flag_na_rule)
            worksheet.conditional_formatting.add(support_flag_area, support_flag_external_rule)
            worksheet.conditional_formatting.add(support_flag_area, support_flag_yes_rule)
            worksheet.conditional_formatting.add(running_area, running_yes_rule)
            worksheet.conditional_formatting.add(running_area, running_no_rule)
            worksheet.conditional_formatting.add(rpm_info_area, no_rpm_rule)

            workbook.save(file)

    def to_pdf(self, driver_tables, file):
        self.to_html(driver_tables, '.tmp.rpms.html')
        pdfkit.from_file('.tmp.rpms.html', file)
        os.remove('.tmp.rpms.html')


    def to_all(self, driver_tables, directory):
        self.to_html(driver_tables, os.path.join(directory, 'check_result.html'))
        self.to_pdf(driver_tables, os.path.join(directory, 'check_result.pdf'))
        self.to_excel(driver_tables, os.path.join(directory, 'check_result.xlsx'))
