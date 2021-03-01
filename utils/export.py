from rich.console import Console
from rich.table import Table
from pathlib import Path
import pdfkit
import pandas as pd
import os
from dominate.tags import body, div, html, h1
from dominate.util import raw
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
import json


def load_export_config():
    with open('config/color.default.json') as jf:
        formatting = json.load(jf)
        return formatting


def load_table_format():
    formatting = load_export_config()
    return formatting["table"]


def load_rpm_info_format():
    formatting = load_export_config()
    return formatting['highlight']['rpm-info']


def load_support_flag_format():
    formatting = load_export_config()
    return formatting['highlight']['support-flag']


def load_running_format():
    formatting = load_export_config()
    return formatting['highlight']['running']


def rpms_export_to_terminal(df):
    console = Console()
    table = Table(show_header=True, header_style='bold green', show_lines=True)
    table.add_column('Name', width=64)
    table.add_column('Vendor', width=12)
    table.add_column('Signature', width=28)
    table.add_column('Distribution', width=12)
    table.add_column('Driver Support Status')

    support_flag_format = load_support_flag_format()

    for _, row in df.iterrows():
        if 'Not supported' in row['Driver Support Status']:
            table.add_row(row['Name'], row['Vendor'], row['Signature'], row['Distribution'], '[red]' + row['Driver Support Status'] + '[/red]')
        elif 'Supported by both SUSE and the vendor' in row['Driver Support Status']:
            table.add_row(row['Name'], row['Vendor'], row['Signature'], row['Distribution'], '[blue]' + row['Driver Support Status'] + '[/blue]')
        elif 'Supported by SUSE' in row['Driver Support Status']:
            table.add_row(row['Name'], row['Vendor'], row['Signature'], row['Distribution'], '[green]' + row['Driver Support Status'] + '[/green]')
        else:
            table.add_row(row['Name'], row['Vendor'], row['Signature'], row['Distribution'], row['Driver Support Status'])

    console.print(table)


def set_rpm_format_for_html(column):
    support_flag_format = load_support_flag_format()

    if 'Not supported' in column:
        return 'background-color: ' + support_flag_format['N/A']['background-color']
    elif 'Supported by both SUSE and the vendor' in column:
        return 'background-color: ' + support_flag_format['external']['background-color']
    elif 'Supported by SUSE' in column:
        return 'background-color: ' + support_flag_format['yes']['background-color']

    return 'background-color: white'


def get_table_render_styles():
    table_formatting = load_table_format()
    styles = [
        dict(selector='table', props=[('border', table_formatting['border'])]),
        dict(selector='th', props=[('font-size', table_formatting['th']['font-size']),
                                   ('font-family', table_formatting['th']['font-family']),
                                   ('background-color', table_formatting['th']['background-color'])]),
        dict(selector='td', props=[('font-size', table_formatting['td']['font-size']),
                                   ('font-family', table_formatting['td']['font-family'])])
    ]

    return styles


def rpms_export_to_html(df, file):
    df['Driver Support Status'] = df['Driver Support Status'].str.replace('\n', '</br>')
    df['Driver Support Status'] = df['Driver Support Status'].str.replace('\t', '&nbsp&nbsp&nbsp&nbsp')
    s = df.style.applymap(set_rpm_format_for_html,
                          subset=pd.IndexSlice[:, ['Driver Support Status']],
                          ).hide_index()

    styles = get_table_render_styles()

    s = s.set_table_styles(styles)

    with open(file, 'w') as f:
        f.write(s.render())


def rpms_export_to_pdf(df, file):
    rpms_export_to_html(df, '.tmp.rpms.html')
    pdfkit.from_file('.tmp.rpms.html', file)
    os.remove('.tmp.rpms.html')


def rpms_export_to_excel(df, file):
    writer = pd.ExcelWriter(file, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Solid Driver Checks')
    workbook = writer.book

    # workbook = load_workbook(file)
    support_flag_format = load_support_flag_format()
    worksheet = writer.sheets['Solid Driver Checks']
    na_format = workbook.add_format({'bg_color': '\'' + support_flag_format['N/A']['background-color'] + '\''})
    yes_format = workbook.add_format({'bg_color': '\'' + support_flag_format['yes']['background-color'] + '\''})
    external_format = workbook.add_format({'bg_color': '\'' + support_flag_format['external']['background-color'] + '\''})

    area = 'E2:E' + str(len(df.index) + 1)
    worksheet.conditional_format(area, {'type': 'text',
                                        'criteria': 'containing',
                                        'value': 'Not supported',
                                        'format': na_format})
    worksheet.conditional_format(area, {'type': 'text',
                                        'criteria': 'containing',
                                        'value': 'Supported by both SUSE and the vendor',
                                        'format': external_format})
    worksheet.conditional_format(area, {'type': 'text',
                                        'criteria': 'containing',
                                        'value': 'Supported by SUSE',
                                        'format': yes_format})

    writer.save()


def rpms_export_to_all(df, directory):
    rpms_export_to_terminal(df)
    rpms_export_to_excel(df, directory + "/check_result.xlsx")
    rpms_export_to_html(df, directory + "/check_result.html")
    rpms_export_to_pdf(df, directory + "/check_result.pdf")


def get_support_flag_with_color_for_rich(support_flag):
    support_flag_format = load_support_flag_format()
    
    if support_flag == 'yes':
        return '[' + support_flag_format['yes']['background-color'] + ']' + 'Supported by SUSE' + '[/' + support_flag_format['yes']['background-color'] + ']'
    elif support_flag == 'external':
        return '[' + support_flag_format['external']['background-color'] + ']' + 'Supported by both SUSE and vendor' + '[/' + support_flag_format['external']['background-color'] + ']'
    else:
        return '[' + support_flag_format['N/A']['background-color'] + ']' + 'N/A' + '[/' + support_flag_format['N/A']['background-color'] + ']'


def get_running_with_color_for_rich(running):
    running_format = load_running_format()
    if running == 'True':
        return '[' + running_format['true']['background-color'] + 'True' + '[/' + running_format['true']['background-color'] + ']'
    else:
        return '[' + running_format['false']['background-color'] + 'False' + '[/' + running_format['false']['background-color'] + ']'


def get_rpm_information_with_color_for_rich(rpm_info):
    rpm_format = load_rpm_info_format()
    if 'is not owned by any package' in rpm_info:
        color = rpm_format['rpm-info']['no-rpm']['background-color']
        return '[' + color + ']' + rpm_info + '[/' + color + ']'
    else:
        return rpm_info


def os_export_to_terminal(df):
    console = Console()
    table = Table(show_header=True, header_style='bold green', show_lines=True)
    table.add_column('Name', width=64)
    table.add_column('Path', width=64)
    table.add_column('Support Flag', width=28)
    table.add_column('Running', width=12)
    table.add_column('RPM Information')

    for _, row in df.iterrows():
        support_flag = get_support_flag_with_color_for_rich(row['Support Flag'])
        running = get_running_with_color_for_rich(row['Running'])
        rpm_info = get_rpm_information_with_color_for_rich(row['RPM Information'])

        table.add_row(row['Name'], row['Path'], support_flag, running, rpm_info)

    console.print(table)


def set_drivers_support_flag_format_for_html(col):
    support_flag_format = load_support_flag_format()
    if col == 'yes':
        return support_flag_format['yes']['background-color']
    elif col == 'external':
        return support_flag_format['external']['background-color']
    elif col == 'N/A':
        return support_flag_format['N/A']['background-color']

    return ''


def set_drivers_running_format_for_html(col):
    running_format = load_running_format()
    if col == 'True':
        return 'background-color:' + running_format['true']['background-color']

    return 'background-color:' + running_format['false']['background-color']


def set_drivers_rpm_info_format_for_html(col):
    rpm_format = load_rpm_info_format()
    if 'is not owned by any package' in col:
        return 'background-color:' + rpm_format['rpm-info']['no-rpm']['background-color']

    return ''


def os_export_to_html(df_collection, file):
    context = html()
    with context:
        with body():
            for key in df_collection:
                h1('Solid Driver Checking Result: ' + key)
                s = df_collection[key].style.applymap(set_drivers_support_flag_format_for_html,
                                                      subset=pd.IndexSlice[:, ['Support Flag']])
                s = s.applymap(set_drivers_running_format_for_html,
                               subset=pd.IndexSlice[:, ['Running']])
                s = s.applymap(set_drivers_rpm_info_format_for_html,
                               subset=pd.IndexSlice[:, ['RPM Information']]).hide_index()

                styles = get_table_render_styles()

                s = s.set_table_styles(styles)
                div(raw(s.render()))

    with open(file, 'w') as f:
        f.write(context.render())

def os_export_to_pdf(df, file):
    rpms_export_to_html(df, '.tmp.os.html')
    pdfkit.from_file('.tmp.os.html', file)
    os.remove('.tmp.os.html')

def os_export_to_excel(df, file, sheet='Solid Driver Checks'):
    writer = pd.ExcelWriter(file, engine='openpyxl')
    if os.path.exists(file):
        writer = pd.ExcelWriter(file, engine='openpyxl', mode='a')

    df.to_excel(writer, index=False, sheet_name=sheet)
    writer.save()
    writer.close()

    workbook = load_workbook(filename=file)
    worksheet = workbook[sheet]

    records = str(len(df.index) + 1)
    support_flag_area = 'C2:C' + records
    running_area = 'D2:D' + records
    rpm_info_area = 'E2:E' + records

    support_flag_format = load_support_flag_format()
    na_text = Font(color=support_flag_format['N/A']['font-color'])
    na_fill = PatternFill(bgColor=support_flag_format['N/A']['background-color'])
    na = DifferentialStyle(font=na_text, fill=na_fill)
    yes_text = Font(color=support_flag_format['yes']['font-color'])
    yes_fill = PatternFill(bgColor=support_flag_format['yes']['background-color'])
    yes = DifferentialStyle(font=yes_text, fill=yes_fill)
    external_text = Font(color=support_flag_format['external']['font-color'])
    external_fill = PatternFill(bgColor=support_flag_format['external']['background-color'])
    external = DifferentialStyle(font=external_text, fill=external_fill)
    running_format = load_running_format()
    false_text = Font(color=running_format['false']['font-color'])
    false_fill = PatternFill(bgColor=running_format['false']['background-color'])
    false_format = DifferentialStyle(font=false_text, fill=false_fill)
    true_text = Font(color=running_format['true']['font-color'])
    true_fill = PatternFill(bgColor=running_format['true']['background-color'])
    true_format = DifferentialStyle(font=true_text, fill=true_fill)
    rpm_info_format = load_rpm_info_format()
    no_rpm_text = Font(color=rpm_info_format['no-rpm']['font-color'])
    no_rpm_fill = PatternFill(bgColor=running_format['no-rpm']['background-color'])
    no_rpm = DifferentialStyle(font=no_rpm_text, fill=no_rpm_fill)
    support_flag_na_rule = Rule(type='containsText', operator='containsText', text='N/A', dxf=na)
    support_flag_external_rule = Rule(type='containsText', operator='containsText', text='external', dxf=external)
    support_flag_yes_rule = Rule(type='containsText', operator='containsText', text='yes', dxf=yes)
    running_yes_rule = Rule(type='containsText', operator='containsText', text='True', dxf=true_format)
    running_no_rule = Rule(type='containsText', operator='containsText', text='False', dxf=false_format)
    no_rpm_rule = Rule(type='containsText', operator='containsText', text='is not owned by any package', dxf=no_rpm)

    worksheet.conditional_formatting.add(support_flag_area, support_flag_na_rule)
    worksheet.conditional_formatting.add(support_flag_area, support_flag_external_rule)
    worksheet.conditional_formatting.add(support_flag_area, support_flag_yes_rule)
    worksheet.conditional_formatting.add(running_area, running_yes_rule)
    worksheet.conditional_formatting.add(running_area, running_no_rule)
    worksheet.conditional_formatting.add(rpm_info_area, no_rpm_rule)

    workbook.save(file)

def os_export_to_all(df, directory):
    os_export_to_terminal(df)
    os_export_to_excel(df, directory + '/check_result.xlsx')
    os_export_to_html(df, directory + '/check_result.html')
    os_export_to_pdf(df, directory + '/check_result.pdf')

def remote_export_to_terminal(driver_collections):
    console = Console()
    for server in driver_collections:
        console.print(server + ' Checking result:')
        os_export_to_terminal(driver_collections[server])


def remote_export_to_html(driver_collections, file):
    os_export_to_html(driver_collections, file)

def remote_export_to_pdf(driver_collections, file):
    remote_export_to_html(driver_collections, '.tmp.remote.html')
    pdfkit.from_file('.tmp.remote.html', file)
    os.remove('.tmp.remote.html')

def remote_export_to_excel(driver_collections, file):
    for server in driver_collections:
        os_export_to_excel(driver_collections[server], file, server)

def remote_export_to_all(driver_collections, directory):
    remote_export_to_terminal(driver_collections)
    remote_export_to_excel(driver_collections, directory + '/check_result.xlsx')
    remote_export_to_html(driver_collections, directory + '/check_result.html')
    remote_export_to_pdf(driver_collections, directory + '/check_result.pdf')

def print_driver(path, driver_support_flag, running, rpm_info):
    console = Console()
    console.print('Name: ' + Path(path).name, style='bold green')
    style = 'bold green'
    if running is False:
        style = 'bold grey85'
    console.print('Running: ', running, style=style)
    if driver_support_flag == 'yes':
        console.print('Support Status: Driver is supported by SUSE', style='bold green')
    elif driver_support_flag == 'external':
        console.print('Support Status: Driver is supported by both SUSE and vendor', style='bold blue')
    elif driver_support_flag == 'N/A':
        console.print("Support Status: Driver don't have support flag!", style='bold red')
    else:
        console.print('Support Status: Unknow, driver support flag is ', driver_support_flag, style='bold red')
    style='bold green'
    if 'is not owned by any package' in rpm_info:
        console.print(rpm_info, style='bold red')
    else:
        console.print('Driver is in rpm: ' + rpm_info, style='bold green')
