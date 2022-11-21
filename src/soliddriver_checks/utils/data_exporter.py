from sqlite3 import Timestamp
import pdfkit
import pandas as pd
import os
from pathlib import Path
from dominate.tags import html, body, h1, div, tr, td, th, table, style, ul, li, p
from dominate.util import raw
from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment,
    NamedStyle,
    borders,
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import string
from openpyxl import Workbook
from jinja2 import Environment, FileSystemLoader
import re
from copy import copy
from datetime import datetime
from ..version import __VERSION__


def _get_version():
    return f"version: {__VERSION__}"

def _generate_timestamp():
    return f"timestamp: {datetime.now()}"


class SDCConf:
    def __init__(self):
        pkg_path = os.path.dirname(__file__)
        cfg_path = f"{pkg_path}/../config/soliddriver-checks.conf"

        with open(cfg_path, "r") as fp:
            self._conf = json.load(fp)

    def _get_xlsx_info(self, *locs):
        conf = self._conf
        for loc in locs:
            conf = conf[loc]

        font = Font(
            name=conf["font"]["family"],
            size=conf["font"]["size"],
            bold=conf["font"]["bold"],
            color=conf["font"]["color"],
        )
        sd = Side(
            border_style=conf["side"]["border_style"], color=conf["side"]["color"]
        )
        bd = Border(top=sd, left=sd, right=sd, bottom=sd)
        fill = PatternFill(
            start_color=conf["fill"]["bgcolor"],
            end_color=conf["fill"]["bgcolor"],
            fill_type="solid",
        )

        return font, bd, fill

    def get_valid_licenses(self):
        return self._conf["valid-licenses"]

    def get_rpm_xslx_table_header(self):
        return self._get_xlsx_info("rpm-check", "excel", "table", "header")

    def get_rpm_xslx_table_normal(self):
        return self._get_xlsx_info("rpm-check", "excel", "table", "data", "normal")

    def get_rpm_xslx_table_important_failed(self):
        return self._get_xlsx_info(
            "rpm-check", "excel", "table", "data", "important-failed"
        )

    def get_rpm_xslx_table_critical_failed(self):
        return self._get_xlsx_info(
            "rpm-check", "excel", "table", "data", "critical-failed"
        )

    def get_rpm_xslx_table_great_row(self):
        return self._get_xlsx_info("rpm-check", "excel", "table", "row", "great")

    def get_rpm_xslx_table_warn_row(self):
        return self._get_xlsx_info("rpm-check", "excel", "table", "row", "warn")

    def get_driver_html_warn_critical(self):
        return self._conf["driver-check"]["html"]["critical_failed"]

    def get_driver_sig_keys(self):
        return self._conf["driver-check"]["sig-keys"]

    def get_driver_html_warn_important(self):
        return self._conf["driver-check"]["html"]["important_failed"]

    def get_driver_xslx_table_header(self):
        return self._get_xlsx_info("driver-check", "excel", "table", "header")

    def get_driver_xslx_table_normal(self):
        return self._get_xlsx_info("driver-check", "excel", "table", "data", "normal")

    def get_driver_xslx_table_important_failed(self):
        return self._get_xlsx_info(
            "driver-check", "excel", "table", "data", "important-failed"
        )

    def get_driver_xslx_table_critical_failed(self):
        return self._get_xlsx_info(
            "driver-check", "excel", "table", "data", "critical-failed"
        )

    def get_driver_xslx_table_warn_row(self):
        return self._get_xlsx_info("driver-check", "excel", "table", "row", "warn")


def ValidLicense(license, licenses):
    for lic in licenses:
        if lic["name"] == license:
            return True

    return False


class ExcelTemplate:
    def __init__(self):
        pkg_path = os.path.dirname(__file__)
        self._cfg_path = f"{pkg_path}/../config/templates/templates.xlsx"

    def set_driver_check_overview(self, ws):
        self._copy_work_sheep("km-report-overview", ws)

    def set_rpm_check_overview(self, ws):
        self._copy_work_sheep("kmp-report-overview", ws)

    def _copy_work_sheep(self, title, ws):
        tmpl = load_workbook(filename=self._cfg_path)
        cover = tmpl[title]

        ws.title = "Overview"
        for row in cover.rows:
            for cell in row:
                ws[cell.coordinate].value = copy(cell.value)
                # ws[cell.coordinate].style = copy(cell.style)
                ws[cell.coordinate].font = copy(cell.font)
                ws[cell.coordinate].border = copy(cell.border)
                ws[cell.coordinate].fill = copy(cell.fill)
                ws[cell.coordinate].alignment = copy(cell.alignment)

        ws["A5"].value = _get_version()
        ws["A6"].value = _generate_timestamp()
        ws.column_dimensions['A'].width = 200

class RPMsExporter:
    def __init__(self):
        self._style = SDCConf()

    def _summary_symbol_result(self, val):
        unfound_no = len(val["unfound"])
        cs_mm_no = len(val["checksum-mismatch"])
        if unfound_no == 0 and cs_mm_no == 0:
            return ""

        result = ""
        if unfound_no > 0:
            result = f"Number of symbols can not be found in KMP: {unfound_no}"

        if cs_mm_no > 0:
            result = result + f"  Number of symbols checksum does not match: {cs_mm_no}"

        return result

    def _get_sym_check_failed(self, val):
        result = ""
        for driver in val:
            summary = self._summary_symbol_result(val[driver])
            if summary != "":
                result = result + "Found {} has below issue(s):\n  {}\n".format(
                    Path(driver).name, summary
                )

        return result

    def _get_supported_driver_failed(self, drivers):
        failed_drivers = []
        for d in drivers:
            if drivers[d] != "external":
                d_name = Path(d).name
                errDetail = f"value is '{drivers[d]}'"
                if drivers[d] == "Missing":
                    errDetail = "no 'supported' flag found"
                errMsg = f"'supported' flag check failed, {errDetail}"
                failed_drivers.append(f"{d_name} : {errMsg}")

        return failed_drivers

    def _get_no_signed_driver(self, drivers):
        results = []
        for d in drivers:
            if not drivers[d]:
                d_name = Path(d).name
                results.append(f"{d_name} : no signature found")

        return results

    def _get_summary_table(self, rpm_table):
        df = rpm_table.copy()

        df["sym-check"] = df["sym-check"].astype(str)

        df.loc[df["sym-check"].str.contains("ko"), "sym-check"] = "failed"

        vendors = df["vendor"].unique()
        df_summary = pd.DataFrame(
            columns=[
                "Vendor",
                "Total KMPs",
                "License",
                "Signature",
                "Weak Module Invoked",
                "Supported Flag/Signature Checks Failed",
                "Symbols Check Failed",
                "Modalias Check Failed", 
            ]
        )

        def alias_check(alias):
            num = 0

            for i in range(len(alias)):
                if alias.iat[i]["match_all"] or len(alias.iat[i]["unmatched_km_alias"]) > 0 or len(alias.iat[i]["unmatched_kmp_alias"]) > 0:
                    num += 1

            return num

        def supported_sig_check(rpms_sf, rpms_is):
            """
            rpms_sf: RPMs' driver supported flag
            rpm_is : RPMs' is_signed column
            """
            num = 0
            # rpms_sf and rpms_is have the same drivers.
            rpms = len(rpms_sf.index)
            for i in range(rpms):
                sf_drivers = rpms_sf.iat[i]
                is_drivers = rpms_is.iat[i]
                found_issues = False
                for key in sf_drivers:
                    if sf_drivers[key] != "external" or not is_drivers[key]:
                        found_issues = True
                        break
                if not found_issues:
                    num += 1

            return rpms - num

        def license_check(vld_lics, rpm_licenses, driver_license):
            count = 0
            for idx, rl in rpm_licenses.items():
                if ValidLicense(rl, vld_lics):
                    count += 1
                else:
                    d_lics = driver_license[idx]
                    for dl in d_lics:
                        if ValidLicense(dl, vld_lics):
                            count += 1
                            break

            return count

        vld_lic = self._style.get_valid_licenses()

        for v in vendors:
            df_vendor = df.loc[df["vendor"] == v]
            total = len(df_vendor.index)
            spc = supported_sig_check(df_vendor["df-supported"], df_vendor["is-signed"])
            symc = len(
                df_vendor.loc[df_vendor["sym-check"] == "failed", "sym-check"].index
            )
            lic_check = license_check(
                vld_lic, df_vendor["license"], df_vendor["dv-licenses"]
            )
            no_sig = len(
                df_vendor.loc[
                    (df_vendor["signature"] != "")
                    & (df_vendor["signature"] != "(none)"),
                    "signature",
                ].index
            )
            wm_invoked = len(df_vendor.loc[df_vendor["wm-invoked"], "wm-invoked"].index)
            aliasc = alias_check(df_vendor["modalias"])

            df_summary = df_summary.append(
                {
                    "Vendor": v,
                    "Total KMPs": total,
                    "License": f"{lic_check} ({lic_check/total * 100:.2f}%)",
                    "Signature": f"{no_sig} ({no_sig/total * 100:.2f}%)",
                    "Weak Module Invoked": f"{wm_invoked} ({wm_invoked/total * 100:.2f}%)",
                    "Supported Flag/Signature Checks Failed": f"{spc} ({spc/total * 100:.2f}%)",
                    "Symbols Check Failed": f"{symc} ({symc/total * 100:.2f}%)",
                    "Modalias Check Failed": f"{aliasc} ({aliasc/total * 100:.2f}%)",
                },
                ignore_index=True,
            )

        return df_summary

    def _get_summary_table_html(self, rpm_table):
        tb = table()
        with tb:
            tb.set_attribute("class", "summary_table")
            df_summary = self._get_summary_table(rpm_table)
            with tr():
                cols = df_summary.columns
                for col in cols:
                    if col == "vendor":
                        t = th(col)
                        t.set_attribute("class", "summary_vendor")
                    else:
                        th(col)

                for i, row in df_summary.iterrows():
                    vendor = row["Vendor"]
                    total_rpms = row["Total KMPs"]
                    ssc = row["Supported Flag/Signature Checks Failed"]
                    lic_check = row["License"]
                    signature = row["Signature"]
                    wm_invoked = row["Weak Module Invoked"]
                    sym_failed = row["Symbols Check Failed"]
                    alias_failed = row["Modalias Check Failed"]

                    row_passed = False
                    if (
                        vendor != ""
                        and int(ssc.split(" ")[0]) == 0
                        and int(lic_check.split(" ")[0]) == total_rpms
                        and int(signature.split(" ")[0]) == total_rpms
                        and int(wm_invoked.split(" ")[0]) == total_rpms
                        and int(sym_failed.split(" ")[0]) == 0
                        and int(alias_failed.split(" ")[0]) == 0
                    ):
                        row_passed = True
                    with tr() as r:
                        if row_passed:
                            r.set_attribute("class", "summary_great_row")
                        if vendor != "":
                            td(vendor)
                        else:
                            tv = td("no vendor information")
                            tv.set_attribute("class", "important_failed")
                        with td(total_rpms) as t:
                            t.set_attribute("class", "summary_total")
                        with td(lic_check) as t:
                            if int(lic_check.split(" ")[0]) != total_rpms:
                                t.set_attribute(
                                    "class", "important_failed summary_number"
                                )
                            else:
                                t.set_attribute("class", "summary_number")
                        with td(signature) as t:
                            if int(signature.split(" ")[0]) != total_rpms:
                                t.set_attribute(
                                    "class", "important_failed summary_number"
                                )
                            else:
                                t.set_attribute("class", "summary_number")
                        with td(wm_invoked) as t:
                            if int(wm_invoked.split(" ")[0]) != total_rpms:
                                t.set_attribute(
                                    "class", "critical_failed summary_number"
                                )
                            else:
                                t.set_attribute("class", "summary_number")
                        with td(ssc) as t:
                            if int(ssc.split(" ")[0]) != 0:
                                t.set_attribute(
                                    "class", "critical_failed summary_number"
                                )
                            else:
                                t.set_attribute("class", "summary_number")
                        with td(sym_failed) as t:
                            if int(sym_failed.split(" ")[0]) != 0:
                                t.set_attribute(
                                    "class", "critical_failed summary_number"
                                )
                            else:
                                t.set_attribute("class", "summary_number")
                        with td(alias_failed) as t:
                            if int(alias_failed.split(" ")[0]) > 0:
                                t.set_attribute(
                                    "class", "important_falied summary_number"
                                )
                            else:
                                t.set_attribute("class", "summary_number")

        return tb

    def _rename_rpm_detail_columns(self, rpm_table):
        df = rpm_table.copy()
        df = df.rename(
            columns={
                "name": "Name",
                "path": "Path",
                "vendor": "Vendor",
                "signature": "Signature",
                "license": "License",
                "wm-invoked": "Weak Module Invoked",
                "df-supported": "Supported Flag Check",
                "sym-check": "Symbols Check",
                "modalias": "Modalias Check",
                "dv-licenses": "Driver Licenses",
                "is-signed": "is-signed",
            }
        )

        return df

    def _fmt_driver_license_check(self, rpm_license, driver_licenses, vld_lics):
        chk_result = ""
        if not ValidLicense(rpm_license, vld_lics):
            chk_result = f"RPM license doesn't supported: {rpm_license}"
            return chk_result

        un_supported_driver = dict()
        for key in driver_licenses:
            if not ValidLicense(driver_licenses[key], vld_lics):
                un_supported_driver[key] = driver_licenses[key]

        if len(un_supported_driver) == 0:
            return chk_result
        else:
            chk_result = "Driver licenses are not supported!\n"
            for idx, key in enumerate(un_supported_driver):
                if idx > 2:  # only show 3 result is enough, keep the table clear.
                    chk_result = f"{chk_result} ..."
                    break
                chk_result = (
                    f"{chk_result} {Path(key).name} : {un_supported_driver[key]}\n"
                )

        return chk_result
    
    def _fmt_modalias_check(self, alias):
        match_all = alias["match_all"]
        km_unmatched = len(alias["unmatched_km_alias"])
        kmp_unmatched = len(alias["unmatched_kmp_alias"])
        
        if match_all:
            return "KMP can match all the devices! Highly not recommended!"
        
        message = ""
        
        if km_unmatched > 0:
            message += "Alias found in kernel module but no match in it's package: "
            for kmu in alias["unmatched_km_alias"]:
                message += kmu + ", "
            message += "\n"
        
        if kmp_unmatched > 0:
            message += "Alias found in the package but no match in it's kernel module: "
            for kmpu in alias["unmatched_kmp_alias"]:
                message += kmpu + ", "
        
        return message
        

    def _get_table_detail_html(self, rpm_table):
        df = self._rename_rpm_detail_columns(rpm_table)
        tb = table()
        vld_lic = self._style.get_valid_licenses()
        with tb:
            tb.set_attribute("class", "table_center")
            with tr():
                th("KMP Checks", colspan=6).set_attribute("class", f"detail_rpm")
                th("Kernel Module Checks", colspan=3).set_attribute("class", f"detail_kernel_module")
            with tr():
                th("Name").set_attribute("class", f"detail_0")
                th("Path").set_attribute("class", f"detail_1")
                th("Vendor").set_attribute("class", f"detail_2")
                th(raw("Signature<span class=\"tooltiptext\">Only check there's a signature or not.</span>")).set_attribute("class", f"detail_3 tooltip")
                th(raw("License<span class=\"tooltiptext\">KMP and it's kernel modules should use open source licenses.</span>")).set_attribute("class", f"detail_4 tooltip")
                th(raw("Weak Module Invoked<span class=\"tooltiptext\">Weak Module is necessary to make 3rd party kernel modules installed for one kernel available to KABI-compatible kernels. </span>")).set_attribute("class", f"detail_5 tooltip")
                th(raw("Supported Flag/Signature<span class=\"tooltiptext\">\"supported\" flag: <br/>  \"yes\": Only supported by SUSE<br/>  \"external\": supported by both SUSE and vendor</span>")).set_attribute("class", f"detail_6 tooltip")
                th(raw("Symbols<span class=\"tooltiptext\">Symbols check is to check whether the symbols in kernel modules matches the symbols in its package.</span>")).set_attribute("class", f"detail_7 tooltip")
                th(raw("Modalias<span class=\"tooltiptext\">Modalias check is to check whether the modalias in kernel modules matches the modalias in its package.</span>")).set_attribute("class", f"detail_8 tooltip")

            for i, row in df.iterrows():
                with tr() as r:
                    name = row["Name"]
                    path = row["Path"]
                    vendor = row["Vendor"]
                    signature = row["Signature"]
                    # distribution = row["Distribution"]
                    license = row["License"]
                    wm_invoked = row["Weak Module Invoked"]
                    sym_check = self._get_sym_check_failed(
                        row["Symbols Check"]
                    ).replace("\n", "</br>")
                    supported = row["Supported Flag Check"]
                    is_signed = row["is-signed"]
                    alias = row["Modalias Check"]
                    alias_chk = self._fmt_modalias_check(alias)
                    dc_err = self._supported_sig_errs(supported, is_signed)
                    dv_license = row["Driver Licenses"]
                    lcs_chk = self._fmt_driver_license_check(
                        license, dv_license, vld_lic
                    )
                    lcs_chk.replace("\n", "</br>")
                    if len(dc_err) > 0:
                        r.set_attribute("class", "critical_failed_row")
                    td(name)
                    td(path)
                    if vendor != "":  # vendor check
                        td(vendor)
                    else:
                        tv = td("no vendor information")
                        tv.set_attribute("class", "important_failed")
                        r.set_attribute("class", "important_failed_row")
                    if signature != "" and signature != "(none)":  # signature check
                        td(signature)
                    else:
                        ts = td(signature)
                        ts.set_attribute("class", "important_failed")
                    if lcs_chk == "":  # license check
                        if license == "":
                            tl = td("No License")
                            tl.set_attribute("class", "important_failed")
                            r.set_attribute("class", "important_failed_row")
                        elif ValidLicense(license, vld_lic):
                            td(license)
                        else:
                            tl = td(license)
                            tl.set_attribute("class", "important_failed")
                            r.set_attribute("class", "important_failed_row")
                    else:
                        tl = td(raw(lcs_chk))
                        tl.set_attribute("class", "important_failed")
                        r.set_attribute("class", "important_failed_row")
                    if wm_invoked or "debuginfo" in name:  # wm check
                        td(str(wm_invoked))
                    else:
                        tw = td(str(wm_invoked))
                        tw.set_attribute("class", "critical_failed")
                        r.set_attribute("class", "critical_failed_row")
                    if len(dc_err) > 0:  # driver check.
                        t_w = td(raw("</br>".join(dc_err)))
                        t_w.set_attribute("class", "critical_failed")
                    else:
                        t = td("All passed!")
                        t.set_attribute("class", "detail_pass")
                    if sym_check == "":  # symbol check.
                        t = td("All passed!")
                        t.set_attribute("class", "detail_pass")
                    else:
                        t_w = td(raw(sym_check))
                        t_w.set_attribute("class", "critical_failed")
                        r.set_attribute("class", "critical_failed_row")
                    if alias_chk == "":
                        t = td("All passed!")
                        t.set_attribute("class", "detail_pass")
                    else:
                        t_w = td(raw(alias_chk.replace("\n", "</br>")))
                        t_w.set_attribute("class", "important_failed")
                        r.set_attribute("class", "important_failed_row")

        return tb

    def _supported_sig_errs(self, sffds, nsds):
        sfds_l = self._get_supported_driver_failed(sffds)
        nsds_l = self._get_no_signed_driver(nsds)

        return sfds_l + nsds_l

    def to_html(self, rpm_table, file):
        pkg_path = os.path.dirname(__file__)
        jinja_tmpl = f"{pkg_path}/../config/templates"
        file_loader = FileSystemLoader(jinja_tmpl)
        env = Environment(loader=file_loader)

        rpm_tmpl = env.get_template("kmp-checks.html.jinja")

        rpm_checks = rpm_tmpl.render(version=_get_version(), timestamp=_generate_timestamp(),
            summary_table=self._get_summary_table_html(rpm_table),
            rpm_details=self._get_table_detail_html(rpm_table),
        )

        with open(file, "w") as f:
            f.write(rpm_checks)

    def to_json(self, rpm_table, file):
        rpm_table.to_json(file, orient="records")

    def _xlsx_create_overview(self, wb):
        et = ExcelTemplate()
        et.set_rpm_check_overview(wb.active)

    def _get_important_failed_style(self):
        (
            ipt_font,
            ipt_border,
            ipt_fill,
        ) = self._style.get_rpm_xslx_table_important_failed()
        return DifferentialStyle(
            font=ipt_font,
            border=ipt_border,
            fill=ipt_fill,
        )

    def _get_critical_failed_style(self):
        (
            ctc_font,
            ctc_border,
            ctc_fill,
        ) = self._style.get_rpm_xslx_table_critical_failed()
        return DifferentialStyle(
            font=ctc_font,
            border=ctc_border,
            fill=ctc_fill,
        )

    def _get_header_style(self):
        (
            header_font,
            header_border,
            header_fill,
        ) = self._style.get_rpm_xslx_table_header()

        return DifferentialStyle(
            font=header_font,
            border=header_border,
            fill=header_fill,
        )

    def _xlsx_create_vendor_summary(self, wb, rpm_table):
        ws_vs = wb.create_sheet("vendor summary")
        sm_table = self._get_summary_table(rpm_table)
        for row in dataframe_to_rows(sm_table, index=False, header=True):
            ws_vs.append(row)

        (
            header_font,
            header_border,
            header_fill,
        ) = self._style.get_rpm_xslx_table_header()

        for cell in ws_vs[1]:
            cell.font = header_font
            cell.border = header_border
            cell.fill = header_fill

        for row in ws_vs[f"A1:G{len(sm_table.index)+1}"]:
            for cell in row:
                cell.border = header_border

        last_record_row_no = len(sm_table.index) + 1
        (
            great_font,
            great_border,
            great_fill,
        ) = self._style.get_rpm_xslx_table_great_row()
        great_row_style = DifferentialStyle(
            font=great_font,
            border=great_border,
            fill=great_fill,
        )
        great_row = Rule(type="expression", dxf=great_row_style)
        great_row.formula = [
            'AND($A2 <> "", VALUE(LEFT($C2, FIND(" ",$C2)-1))=$B2, VALUE(LEFT($D2, FIND(" ", $D2) - 1)) = $B2, VALUE(LEFT($E2, FIND(" ", $E2) - 1)) = $B2, VALUE(LEFT($F2, FIND(" ", $F2)-1))=$B2, VALUE(LEFT($G2, FIND(" ", $G2) - 1))=0)'
        ]
        ws_vs.conditional_formatting.add(f"A2:G{last_record_row_no}", great_row)

        ipt_style = self._get_important_failed_style()
        ctc_style = self._get_critical_failed_style()

        empty_vendor = Rule(type="expression", dxf=ipt_style)
        empty_vendor.formula = ['$A2 = ""']
        ws_vs.conditional_formatting.add(f"A2:A{last_record_row_no}", empty_vendor)

        supported_failed = Rule(type="expression", dxf=ctc_style)
        supported_failed.formula = ['VALUE(LEFT($C2, FIND(" ", $C2) - 1)) <> $B2']
        ws_vs.conditional_formatting.add(f"C2:C{last_record_row_no}", supported_failed)

        license_check = Rule(type="expression", dxf=ipt_style)
        license_check.formula = ['VALUE(LEFT($D2, FIND(" ", $D2) - 1)) <> $B2']
        ws_vs.conditional_formatting.add(f"D2:D{last_record_row_no}", license_check)

        sig_check = Rule(type="expression", dxf=ipt_style)
        sig_check.formula = ['VALUE(LEFT($E2, FIND(" ", $E2) - 1)) <> $B2']
        ws_vs.conditional_formatting.add(f"E2:E{last_record_row_no}", sig_check)

        wm_check = Rule(type="expression", dxf=ipt_style)
        wm_check.formula = ['VALUE(LEFT($F2, FIND(" ", $F2) - 1)) <> $B2']
        ws_vs.conditional_formatting.add(f"F2:F{last_record_row_no}", wm_check)

        sym_failed = Rule(type="expression", dxf=ctc_style)
        sym_failed.formula = ['VALUE(LEFT($G2, FIND(" ", $G2) - 1)) <> 0']
        ws_vs.conditional_formatting.add(f"G2:G{last_record_row_no}", sym_failed)
        
        ws_vs.column_dimensions['A'].width = 30
        ws_vs.column_dimensions['B'].width = 15
        ws_vs.column_dimensions['C'].width = 15
        ws_vs.column_dimensions['D'].width = 15
        ws_vs.column_dimensions['E'].width = 15
        ws_vs.column_dimensions['F'].width = 15
        ws_vs.column_dimensions['G'].width = 15
        ws_vs.column_dimensions['H'].width = 15
        
    def _xlsx_create_rpm_details(self, wb, rpm_table):
        df = self._rename_rpm_detail_columns(rpm_table)
        ws_rd = wb.create_sheet("KMPs details")
        (
            normal_font,
            normal_border,
            normal_fill,
        ) = self._style.get_rpm_xslx_table_normal()
        normal_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        (
            imt_font,
            imt_border,
            imt_fill,
        ) = self._style.get_rpm_xslx_table_important_failed()
        (
            ctc_font,
            ctc_border,
            ctc_fill,
        ) = self._style.get_rpm_xslx_table_critical_failed()
        ctc_style = self._get_critical_failed_style()
        imt_style = self._get_important_failed_style()
        (
            header_font,
            header_border,
            header_fill,
        ) = self._style.get_rpm_xslx_table_header()

        # Add headers
        ws_rd.merge_cells('A1:G1')
        ws_rd['A1'] = "KMP Checks"
        ws_rd['A1'].font = header_font
        ws_rd['A1'].fill = header_fill
        ws_rd['A1'].border = header_border
        ws_rd['A1'].alignment = center_align
        
        ws_rd.merge_cells('H1:L1')
        ws_rd['H1'] = "Kernel Module Checks"
        ws_rd['H1'].font = header_font
        ws_rd['H1'].fill = header_fill
        ws_rd['H1'].border = header_border
        ws_rd['H1'].alignment = center_align
        
        cols = df.columns
        xlsx_cols = list(string.ascii_lowercase[0:len(cols)])
        for i in range(len(xlsx_cols)):
            cell_no = xlsx_cols[i] + "2"
            ws_rd[cell_no] = cols[i]
            ws_rd[cell_no].font = header_font
            ws_rd[cell_no].fill = header_fill
            ws_rd[cell_no].border = header_border

        ws_rd["H2"] = "Supported Flag/Signature"  # rename the column name
        # No need to show below volumns
        ws_rd.column_dimensions["J"].hidden = True  # Driver Licenses
        ws_rd.column_dimensions["E"].hidden = True  # Distribution
        ws_rd.column_dimensions["K"].hidden = True  # is-signed

        vld_lic = self._style.get_valid_licenses()

        for i, row in df.iterrows():
            rpm_license = row["License"]
            for col_idx in range(len(cols)):
                cell_no = xlsx_cols[col_idx] + str(i+3)
                val = row[cols[col_idx]]

                ws_rd[cell_no].font = normal_font
                ws_rd[cell_no].fill = normal_fill
                ws_rd[cell_no].border = normal_border
                ws_rd[cell_no].alignment = normal_align

                if cols[col_idx] == "Symbols Check":
                    val = self._get_sym_check_failed(val)
                    if val == "":
                        val = "All passed!"
                        ws_rd[cell_no] = val
                        ws_rd[cell_no].alignment = center_align
                    else:
                        ws_rd[cell_no] = val
                        ws_rd[cell_no].font = ctc_font
                        ws_rd[cell_no].fill = ctc_fill
                        ws_rd[cell_no].border = ctc_border
                elif cols[col_idx] == "Supported Flag Check":
                    val = self._supported_sig_errs(val, row["is-signed"])
                    if len(val) > 0:
                        ws_rd[cell_no] = '\n'.join(val)
                        ws_rd[cell_no].font = ctc_font
                        ws_rd[cell_no].fill = ctc_fill
                        ws_rd[cell_no].border = ctc_border
                    else:
                        ws_rd[cell_no] = "All passed!"

                    ws_rd[cell_no].alignment = center_align
                elif cols[col_idx] == "License":
                    lcs_chk = self._fmt_driver_license_check(
                        rpm_license, row["Driver Licenses"], vld_lic
                    )
                    if lcs_chk == "":
                        if rpm_license == "":
                            ws_rd[cell_no] = "No License"
                            ws_rd[cell_no].font = imt_font
                            ws_rd[cell_no].fill = imt_fill
                            ws_rd[cell_no].border = imt_border
                        else:
                            ws_rd[cell_no] = rpm_license
                    else:
                        ws_rd[cell_no] = lcs_chk
                        ws_rd[cell_no].font = imt_font
                        ws_rd[cell_no].fill = imt_fill
                        ws_rd[cell_no].border = imt_border
                elif cols[col_idx] == "Modalias Check":
                    alias_check = self._fmt_modalias_check(row["Modalias Check"])
                    if alias_check != "":
                        ws_rd[cell_no] = alias_check
                        ws_rd[cell_no].font = imt_font
                        ws_rd[cell_no].fill = imt_fill
                        ws_rd[cell_no].border = imt_border
                    else:
                        val = "All passed!"
                        ws_rd[cell_no] = val
                        ws_rd[cell_no].alignment = center_align
                    
                else:  # no format needed.
                    ws_rd[cell_no] = str(val)

        records = len(df) + 1

        # vendor check format
        epty_vendor = Rule(type="expression", dxf=imt_style)
        epty_vendor.formula = ['$C2 = ""']
        ws_rd.conditional_formatting.add(f"C2:C{records}", epty_vendor)

        # signature check format
        sig_rule = Rule(type="expression", dxf=ctc_style)
        sig_rule.formula = ['=OR($D2 = "", $D2 = "(none)")']
        ws_rd.conditional_formatting.add(f"D2:D{records}", sig_rule)
        
        ws_rd.column_dimensions['A'].width = 30
        ws_rd.column_dimensions['B'].width = 40
        ws_rd.column_dimensions['C'].width = 30
        ws_rd.column_dimensions['D'].width = 30
        ws_rd.column_dimensions['F'].width = 20
        ws_rd.column_dimensions['G'].width = 15
        ws_rd.column_dimensions['H'].width = 40
        ws_rd.column_dimensions['I'].width = 40
        ws_rd.column_dimensions['L'].width = 40

    def _xlsx_create_report_workbook(self):
        wb = Workbook()

        return wb

    def to_excel(self, rpm_table, file):
        if os.path.exists(file):
            os.remove(file)

        wb = self._xlsx_create_report_workbook()

        self._xlsx_create_overview(wb)
        self._xlsx_create_vendor_summary(wb, rpm_table)
        self._xlsx_create_rpm_details(wb, rpm_table)

        wb.save(file)

    def to_pdf(self, rpm_table, file):
        self.to_html(rpm_table, ".tmp.kmps.html")
        pdfkit.from_file(".tmp.kmps.html", file)
        os.remove(".tmp.kmps.html")

    def to_all(self, rpm_table, directory):
        excel_file = os.path.join(directory, "check_report.xlsx")
        html_file = os.path.join(directory, "check_report.html")
        pdf_file = os.path.join(directory, "check_report.pdf")
        json_file = os.path.join(directory, "check_report.json")

        if not os.path.exists(directory):
            os.mkdir(directory)
        else:
            if os.path.exists(excel_file):
                os.remove(excel_file)
            if os.path.exists(html_file):
                os.remove(html_file)
            if os.path.exists(pdf_file):
                os.remove(pdf_file)
            if os.path.exists(json_file):
                os.remove(json_file)

        self.to_excel(rpm_table, excel_file)
        self.to_html(rpm_table, html_file)
        self.to_pdf(rpm_table, pdf_file)
        self.to_json(rpm_table, json_file)


class DriversExporter:
    def __init__(self):
        self._style = SDCConf()

    def _driver_path_check(self, driver_path):
        # if the driver is located in weak-updates directory,
        # check wether it's valid or not first.
        for i, wud_checks in self._wu_dt.iterrows():
            if driver_path == wud_checks["driver"] and wud_checks["result"] == "Failed":
                return None

        # TODO: this should be optimized!
        result = re.match(
            r"^/lib/modules/[0-9]+.[0-9]+.[0-9]+\-[0-9]+\-[a-z]+/(updates/|weak-updates/|extra/)",
            driver_path,
        )

        if result is not None:
            return result
        else:
            return re.match(
                r"^/lib/modules/[0-9]+.[0-9]+.[0-9]+\-[0-9]+.[0-9]+\-[a-z]+/(updates/|weak-updates/|extra/)",
                driver_path,
            )

    def _format_row_html(self, row):
        path = row["path"]
        supported = row["flag_supported"]
        supported = supported.split(" ")
        license = row["license"]
        signature = row["signature"]
        rpm = row["rpm"]
        vld_lic = self._style.get_valid_licenses()

        critical_style = self._style.get_driver_html_warn_critical()
        important_style = self._style.get_driver_html_warn_important()
        cs_bgColor = critical_style["background-color"]
        cs_color = critical_style["color"]
        cs_border = critical_style["border"]

        is_bgColor = important_style["background-color"]
        is_border = important_style["border"]

        warn_level = 0  # 0: normal, 1: important, 2: critical

        name_style = ""
        path_style = ""
        supported_style = ""
        license_style = ""
        signature_style = ""
        suse_release_style = ""
        running_style = ""
        rpm_style = ""

        if not ValidLicense(license, vld_lic):
            warn_level = 1
            license_style = f"background-color:{is_bgColor}"

        if not signature:
            warn_level = 1
            signature_style = f"background-color:{is_bgColor}"

        if not self._driver_path_check(path):
            warn_level = 2
            path_style = f"background-color:{cs_bgColor} color:{cs_color}"

        if (len(supported) == 0) or (
            len(supported) == 1 and supported[0] != "external"
        ):
            warn_level = 2
            supported_style = f"background-color:{cs_bgColor} color:{cs_color}"
        if len(supported) > 1:
            v1 = supported[0]
            if v1 != "external":
                warn_level = 2
                supported_style = f"background-color:{cs_bgColor} color:{cs_color}"
            else:
                warn_level = 1
                supported_style = f"background-color:{is_bgColor}"
                for v in supported:
                    if v != v1 and v != "external":
                        warn_level = 2
                        supported_style = (
                            f"background-color:{cs_bgColor} color:{cs_color}"
                        )

        if "is not owned by any package" in rpm:
            warn_level = 2
            rpm_style = f"background-color:{cs_bgColor} color:{cs_color}"

        row_style = [
            name_style,
            path_style,
            supported_style,
            license_style,
            signature_style,
            suse_release_style,
            running_style,
            rpm_style,
        ]
        return row_style

    def to_json(self, driver_tables, file):
        jf = dict()
        for label, driver_table in driver_tables.items():
            if driver_table is not None:
                drivers = driver_table["drivers"]
                wu_drivers = driver_table["weak-update-drivers"]
                drivers_buff = drivers.to_json(orient="records")
                wu_drviers_buff = wu_drivers.to_json(orient="records")
                noinfo_drivers = json.dumps(driver_table["noinfo-drivers"])
            else:
                drivers_buff = "{}"
                wu_drviers_buff = "{}"

            jf[label] = {
                "drivers": json.loads(drivers_buff),
                "weak-update-drivers": json.loads(wu_drviers_buff),
                "noinfo-drivers": json.loads(noinfo_drivers),
            }

        with open(file, "w") as fp:
            json.dump(jf, fp)

    def _get_third_party_drivers(self, drivers):
        df = drivers.copy()
        sig_keys = self._style.get_driver_sig_keys()
        keys = []
        for k in sig_keys:
            keys.append(k["key"])

        ser = ~df.rpm_sig_key.isin(keys)
        df = df[ser]
        df.drop("rpm_sig_key", axis=1, inplace=True)

        return df

    def _refmt_supported(self, drivers):
        df = drivers.copy()
        for i, row in df.iterrows():
            row["flag_supported"] = " ".join(row["flag_supported"])

        return df

    def _append_noinfo_drivers(self, noinfo_drivers, drivers):
        for driver in noinfo_drivers:
            row = [
                driver,
                "Can not find file under /lib/modules",
                "",
                "",
                "",
                "",
                "True",
                f"driver {driver} is not owned by any package",
            ]
            drivers = drivers.append(
                pd.Series(
                    row,
                    index=[
                        "name",
                        "path",
                        "flag_supported",
                        "license",
                        "signature",
                        "os-release",
                        "running",
                        "rpm",
                    ],
                ),
                ignore_index=True,
            )

        return drivers

    def to_html(self, driver_tables, file):
        pkg_path = os.path.dirname(__file__)
        jinja_tmpl = f"{pkg_path}/../config/templates"
        file_loader = FileSystemLoader(jinja_tmpl)
        env = Environment(loader=file_loader)

        driver_tmpl = env.get_template("driver-checks.html.jinja")

        details = []
        for label, driver_table in driver_tables.items():
            if driver_table is None:
                details.append({"name": label, "table": "Connect error!"})
                continue

            dt = driver_table["drivers"]
            self._wu_dt = driver_table["weak-update-drivers"]
            total_drivers, tp_drivers, failed_drivers = self._get_server_summary(dt)
            df = dt.copy()
            df = self._get_third_party_drivers(df)

            # add no information drivers
            noinfo_drivers = driver_table["noinfo-drivers"]
            df = self._append_noinfo_drivers(noinfo_drivers, df)
            df.loc[df["running"] == "True", "running"] = "&#9989;"
            df.loc[df["running"] == "False", "running"] = "&#9940;"
            df = self._refmt_supported(df)
            ts = (
                # df.style.hide(axis='index')
                df.style.hide_index()
                .set_table_attributes('class="table_center"')
                .apply(self._format_row_html, axis=1)
            )

            details.append(
                {
                    "name": label,
                    "total_drivers": total_drivers,
                    "third_party_drivers": tp_drivers,
                    "failed_drivers": failed_drivers,
                    "table": ts.render(),
                }
            )

        driver_checks = driver_tmpl.render(version=_get_version(), timestamp=_generate_timestamp(), details=details)

        with open(file, "w") as f:
            f.write(driver_checks)

    def _xlsx_create_overview(self, wb):
        et = ExcelTemplate()
        et.set_driver_check_overview(wb.active)

    def _get_failed_driver_count(self, df):
        count = 0
        vld_lic = self._style.get_valid_licenses()
        for i, row in df.iterrows():
            if (
                "is not owned by any package" in row["rpm"]
                or " ".join(row["flag_supported"]) != "external"
                or str(row["signature"]) != "True"
                or not self._driver_path_check(row["path"])
                or not ValidLicense(row["license"], vld_lic)
            ):
                count += 1

        return count

    def _get_server_summary(self, driver_table):
        total_drivers = len(driver_table.index)
        third_party_drivers = self._get_third_party_drivers(driver_table)
        tpd_count = len(third_party_drivers.index)

        failed_count = self._get_failed_driver_count(third_party_drivers)

        return total_drivers, tpd_count, failed_count

    def _xlsx_create_table(self, wb, label, driver_tables):
        ws_dc = wb.create_sheet(label)
        if driver_tables is None:
            return

        dt = driver_tables["drivers"]
        self._wu_dt = driver_tables["weak-update-drivers"]
        noinfo_drivers = driver_tables["noinfo-drivers"]
        df = self._get_third_party_drivers(dt)
        noinfo_drivers = driver_tables["noinfo-drivers"]
        df = self._append_noinfo_drivers(noinfo_drivers, df)
        df = self._refmt_supported(df)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws_dc.append(row)

        (
            header_font,
            header_border,
            header_fill,
        ) = self._style.get_driver_xslx_table_header()
        for cell in ws_dc[1]:
            cell.font = header_font
            cell.border = header_border
            cell.fill = header_fill

        (
            ctc_font,
            ctc_border,
            ctc_fill,
        ) = self._style.get_driver_xslx_table_critical_failed()
        (
            imt_font,
            imt_border,
            imt_fill,
        ) = self._style.get_driver_xslx_table_important_failed()
        (
            normal_font,
            normal_border,
            normal_fill,
        ) = self._style.get_driver_xslx_table_normal()
        vld_lic = self._style.get_valid_licenses()

        last_record_row_no = len(df.index) + 2  # the title of the table also count
        for i in range(2, last_record_row_no):
            # default format.
            for cell in ws_dc[i]:
                cell.font = normal_font
                cell.border = normal_border
                cell.fill = normal_fill

            path = ws_dc[f"B{i}"].value
            flag = ws_dc[f"C{i}"].value.split(" ")
            license = ws_dc[f"D{i}"].value
            sig = ws_dc[f"E{i}"].value
            rpm = ws_dc[f"H{i}"].value

            w_level = 0

            if not ValidLicense(license, vld_lic):
                w_level = 1
                ws_dc[f"D{i}"].font = imt_font
                ws_dc[f"D{i}"].border = imt_border
                ws_dc[f"D{i}"].fill = imt_fill

            if sig == "":
                w_level = 1
                ws_dc[f"E{i}"].font = imt_font
                ws_dc[f"E{i}"].border = imt_border
                ws_dc[f"E{i}"].fill = imt_fill

            if not self._driver_path_check(path):
                w_level = 2
                ws_dc[f"B{i}"].font = imt_font
                ws_dc[f"B{i}"].border = imt_border
                ws_dc[f"B{i}"].fill = imt_fill

            if (len(flag) == 0) or (len(flag) == 1 and flag[0] != "external"):
                w_level = 2
                ws_dc[f"C{i}"].font = ctc_font
                ws_dc[f"C{i}"].border = ctc_border
                ws_dc[f"C{i}"].fill = ctc_fill
            elif len(flag) > 1:
                v1 = flag[0]
                if v1 != "external":
                    warn_level = 2
                    ws_dc[f"C{i}"].font = ctc_font
                    ws_dc[f"C{i}"].border = ctc_border
                    ws_dc[f"C{i}"].fill = ctc_fill
                else:
                    warn_level = 1
                    ws_dc[f"C{i}"].font = imt_font
                    ws_dc[f"C{i}"].border = imt_border
                    ws_dc[f"C{i}"].fill = imt_fill
                    for v in flag:
                        if v != v1:
                            warn_level = 2
                            ws_dc[f"C{i}"].font = ctc_font
                            ws_dc[f"C{i}"].border = ctc_border
                            ws_dc[f"C{i}"].fill = ctc_fill
                            break

            if "is not owned by any package" in rpm:
                warn_level = 2
                ws_dc[f"H{i}"].font = ctc_font
                ws_dc[f"H{i}"].border = ctc_border
                ws_dc[f"H{i}"].fill = ctc_fill

    def to_excel(self, driver_tables, file):
        wb = Workbook()
        self._xlsx_create_overview(wb)

        for label, dt in driver_tables.items():
            self._xlsx_create_table(wb, label, dt)

        if os.path.exists(file):
            os.remove(file)

        wb.save(file)

    def to_pdf(self, driver_tables, file):
        self.to_html(driver_tables, ".tmp.rpms.html")

        content = ""
        with open(".tmp.rpms.html", "r") as fp:
            content = fp.read()

        content = str(content).replace("&#9989;", "True").replace("&#9940;", "False")
        with open(".tmp.rpms.html", "w") as fp:
            fp.write(content)

        pdfkit.from_file(".tmp.rpms.html", file)
        os.remove(".tmp.rpms.html")

    def to_all(self, driver_tables, directory):
        excel_file = os.path.join(directory, "check_result.xlsx")
        html_file = os.path.join(directory, "check_result.html")
        pdf_file = os.path.join(directory, "check_result.pdf")
        json_file = os.path.join(directory, "check_result.json")

        if not os.path.exists(directory):
            os.mkdir(directory)
        else:
            if os.path.exists(excel_file):
                os.remove(excel_file)
            if os.path.exists(html_file):
                os.remove(html_file)
            if os.path.exists(pdf_file):
                os.remove(pdf_file)
            if os.path.exists(json_file):
                os.remove(json_file)

        self.to_excel(driver_tables, excel_file)
        self.to_html(driver_tables, html_file)
        self.to_pdf(driver_tables, pdf_file)
        self.to_pdf(driver_tables, json_file)
