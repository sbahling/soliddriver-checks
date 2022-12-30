import os
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from openpyxl import Workbook
from jinja2 import Environment, FileSystemLoader
from ..config import SDCConf, get_version, generate_timestamp
from ..api.km import KMEvaluation
from ..api.analysis import kms_to_dataframe, kms_to_json
from .xlsx_utils import XlsxTemplate, KMXlsxStyler


class KMReporter:
    def __init__(self):
        self._style = SDCConf()

    def _format_cell(slef, value):
        if type(value) == dict:
            return value.get("value", "")
        else:
            return value

    def _row_style_in_html(self, row):
        # TODO: row style should be added
        def _get_cell_style(row_level, cell_level):
            style = ""
            cristyle = self._style.get_km_html_error()
            impstyle = self._style.get_km_html_warning()
            cri_bgcolor = cristyle["background-color"]
            cri_color = cristyle["color"]
            # cri_border = cristyle["border"]

            imp_bgcolor = impstyle["background-color"]
            # imp_border = impstyle["border"]

            if int(KMEvaluation.WARNING) == cell_level:
                style = f"background-color:{imp_bgcolor}"
            elif int(KMEvaluation.ERROR) == cell_level:
                style = f"background-color:{cri_bgcolor} color:{cri_color}"

            return style

        return [
            "",  # level style, no need for this.
            _get_cell_style(KMEvaluation(row["level"]["value"]), KMEvaluation(row["Module Name"]["level"]["value"])),
            _get_cell_style(KMEvaluation(row["level"]["value"]), KMEvaluation(row["File"]["level"]["value"])),
            _get_cell_style(KMEvaluation(row["level"]["value"]), KMEvaluation(row["License"]["level"]["value"])),
            _get_cell_style(KMEvaluation(row["level"]["value"]), KMEvaluation(row["Signature"]["level"]["value"])),
            _get_cell_style(KMEvaluation(row["level"]["value"]), KMEvaluation(row['"supported" Flag']["level"]["value"])),
            "",  # running style, no need for this.
            _get_cell_style(KMEvaluation(row["level"]["value"]), KMEvaluation(row["KMP"]["level"]["value"]))
        ]

    def _format_columns(self, df):
        return df.rename(
            columns = {
                "modulename" : "Module Name",
                "filename"   : "File",
                "license"    : "License",
                "signature"  : "Signature",
                "supported"  : '"supported" Flag',
                "running"    : "Running",
                "kmp"        : "KMP"
            }
        )

    def to_html(self, sys_info, df_format, file):
        pkg_path = os.path.dirname(__file__)
        jinja_tmpl = f"{pkg_path}/../config/templates"
        file_loader = FileSystemLoader(jinja_tmpl)
        env = Environment(loader=file_loader)

        km_tmpl = env.get_template("km-report.html.jinja")

        if df_format is None:
            df_format = kms_to_dataframe()  # read from local system.

        kms_in_total = len(df_format.index)
        failed_kms_in_total = len([f for f in df_format["level"].to_list() if f['value'] != int(KMEvaluation.PASS)])

        df_format = self._format_columns(df_format)
        ts = df_format.style.hide(axis="index").hide([('level')], axis="columns").set_table_attributes('class="table_center"').apply(self._row_style_in_html, axis=1).format(self._format_cell)

        # df["running"].loc[df.running == True] = "&#9989;"
        # df["running"].loc[df.running == False] = "&#9940;"
        # df["running"].loc[df.running == ""] = "N/A"

        # Will be easier to get the value if run this after reformat the values.
        external_kms_in_total = len([sf for sf in df_format['"supported" Flag'].to_list() if sf == "external"])

        kms_buffer = km_tmpl.render(
            version               = get_version(),
            timestamp             = generate_timestamp(),
            sysinfo               = sys_info,
            kms_in_total          = kms_in_total,
            external_kms_in_total = external_kms_in_total,
            failed_kms_in_total   = failed_kms_in_total,
            kms_table             = ts.to_html()
            )

        with open(file, "w") as f:
            f.write(kms_buffer)

    def _create_xlsx_overview(self, ws):
        et = XlsxTemplate()
        et.set_km_overview(ws)

    def _create_xlsx_sheet(self, wb, sys_info, df):
        ws = wb.create_sheet(sys_info)

        df_values = df.copy()
        df_values = self._format_columns(df_values)
        df_values = df_values.drop(['level'], axis=1)
        df_values = df_values.applymap(lambda v: v.get("value", ""))
        # fill the values
        df_values = df_values.astype(str)
        for row in dataframe_to_rows(df_values, index=False, header=True):
            ws.append(row)

        # format table
        render = KMXlsxStyler()
        # format header
        for cell in ws[1]:
            render.set_header(cell)

        # format data
        pair = {'A' : "modulename",
        'B' : "filename",
        'C' : "license",
        'D' : "signature",
        'E' : 'supported',
        'F' : "running",
        'G' : 'kmp'}

        data_start_row = 2
        row_count = len(df.index) + data_start_row
        for i in range(data_start_row, row_count):
            # TODO: add row level style.
            row_level = df.at[i-data_start_row, 'level']
            for cell in ws[i]:
                v = df.at[i-data_start_row, pair[cell.column_letter]]
                lev = v.get('level')
                if lev['value'] == int(KMEvaluation.PASS):
                    render.normal(cell)
                elif lev['value'] == int(KMEvaluation.WARNING):
                    render.warning(cell)
                elif lev['value'] == int(KMEvaluation.ERROR):
                    render.error(cell)

        render.set_column_width(ws, {'A': 25,
                                     'B': 90,
                                     'C': 18,
                                     'D': 10,
                                     'E': 15,
                                     'F': 10,
                                     'G': 30})

    def to_xlsx(self, sys_info, df_format, file):
        wb = Workbook()
        self._create_xlsx_overview(wb.active)

        if df_format is None:
            df_format = kms_to_dataframe()

        self._create_xlsx_sheet(wb, sys_info, df_format)

        wb.save(file)

    def to_json(self, sys_info, df_format, file):
        buffer = kms_to_json(df_format)

        # TODO: add sys_info to json output.
        with open(file, "w") as fp:
            json.dump(json.loads(buffer), fp)
