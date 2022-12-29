import pandas as pd
import os
from dominate.tags import tr, td, th, table
from dominate.util import raw
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from jinja2 import Environment, FileSystemLoader
from ..config import SDCConf, get_version, generate_timestamp
from ..api.kmp import KMPEvaluation
from .xlsx_utils import XlsxTemplate, KMPXlsxStyler


class KMPReporter:
    def __init__(self):
        self._style = SDCConf()

    def _summary(self, df):
        summary = df.copy()
        for i, row in summary.iterrows(): # we need to do this for unique() since unhashable type: 'dict'.
            row["vendor"] = row["vendor"]["value"]

        def failed_len(col):
            counter = 0
            for v in col:
                if v.get("level") != KMPEvaluation.PASS:
                    counter += 1
            return counter

        def format_cell(number, total):
            return f"{number} ({number/total * 100:.2f}%)"

        vendors = summary["vendor"].unique()
        sum_table = pd.DataFrame()
        for v in vendors:
            vendor_df          = summary.loc[summary["vendor"] == v]
            total              = len(vendor_df.index)
            sig_failed         = failed_len(vendor_df["signature"])
            license_failed     = failed_len(vendor_df["license"])
            supported_failed   = failed_len(vendor_df["supported_flag"])
            wm2_invoked_failed = failed_len(vendor_df["wm2_invoked"])
            km_sigs_failed     = failed_len(vendor_df["km_signatures"])
            km_license_failed  = failed_len(vendor_df["km_licenses"])
            symbols_failed     = failed_len(vendor_df["symbols"])
            alias_failed       = failed_len(vendor_df["modalias"])

            new_row = pd.Series({
                "Vendor"             : v,
                "Total KMPs"         : total,
                "License"            : format_cell(license_failed, total),
                "KMP Signature"      : format_cell(sig_failed, total),
                "Weak Module Invoked": format_cell(wm2_invoked_failed, total),
                "Supported Flag"     : format_cell(supported_failed, total),
                "KM Signatures"      : format_cell(km_sigs_failed, total),
                "KM Licenses"        : format_cell(km_license_failed, total),
                "Symbols"            : format_cell(symbols_failed, total),
                "Modalias"           : format_cell(alias_failed, total),
            })

            sum_table = pd.concat([sum_table, new_row.to_frame().T], ignore_index=True)
        
        return sum_table

    def _summary_to_html(self, df):
        tb = table()
        with tb:
            tb.set_attribute("class", "summary_table")
            summary = self._summary(df)
            with tr():
                cols = summary.columns
                for col in cols:
                    if col == "Vendor":
                        t = th(col)
                        t.set_attribute("class", "summary_vendor")
                    else:
                        th(col)
            
            def _pass(item):
                return int(item.split(" ")[0]) == 0
            
            for __, row in summary.iterrows():
                vendor = row["Vendor"]
                total = row["Total KMPs"]
                sig = row["KMP Signature"]
                license = row["License"]
                supported = row["Supported Flag"]
                wm2_invoked = row["Weak Module Invoked"]
                km_sigs = row["KM Signatures"]
                km_license = row["KM Licenses"]
                symbols = row["Symbols"]
                alias = row["Modalias"]
                
                row_passed = False
                if (
                    vendor != ""
                    and _pass(sig)
                    and _pass(license)
                    and _pass(supported)
                    and _pass(wm2_invoked)
                    and _pass(km_sigs)
                    and _pass(km_license)
                    and _pass(symbols)
                    and _pass(alias)
                ):
                    row_passed = True
                with tr() as r:
                    if row_passed:
                        r.set_attribute("class", "summary_great_row")
                    if vendor != "":
                        td(vendor)
                    else:
                        tv = td("no vendor information")
                        tv.set_attribute("class", "important_failed")
                    with td(total) as t:
                        t.set_attribute("class", "summary_total")
                    with td(license) as t:
                        if _pass(license):
                            t.set_attribute("class", "summary_number")
                        else:
                            t.set_attribute("class", "important_failed summary_number")
                    with td(sig) as t:
                        if _pass(sig):
                            t.set_attribute("class", "summary_number")
                        else:
                            t.set_attribute("class", "important_failed summary_number")
                    with td(wm2_invoked) as t:
                        if _pass(wm2_invoked):
                            t.set_attribute("class", "summary_number")
                        else:
                            t.set_attribute("class", "critical_failed summary_number")
                    with td(supported) as t:
                        if _pass(supported):
                            t.set_attribute("class", "summary_number")
                        else:
                            t.set_attribute("class", "critical_failed summary_number")
                    with td(km_sigs) as t:
                        if _pass(km_sigs):
                            t.set_attribute("class", "summary_number")
                        else:
                            t.set_attribute("class", "important_failed summary_number")
                    with td(km_license) as t:
                        if _pass(km_license):
                            t.set_attribute("class", "summary_number")
                        else:
                            t.set_attribute("class", "important_failed summary_number")
                    with td(symbols) as t:
                        if _pass(symbols):
                            t.set_attribute("class", "summary_number")
                        else:
                            t.set_attribute("class", "critical_failed summary_number")
                    with td(alias) as t:
                        if _pass(alias):
                            t.set_attribute("class", "summary_number")
                        else:
                            t.set_attribute("class", "critical_failed summary_number")

        return tb

    def _detail_to_html(self, df):
        def _create_cell(ana_val):
            level, value = ana_val.get("level"), ana_val.get("value")
            if value == None:
                value = ""

            if level == KMPEvaluation.PASS:
                return td(value)
            elif level == KMPEvaluation.WARNING:
                return td(value).set_attribute("class", "important_failed")
            elif level == KMPEvaluation.ERROR:
                return td(value).set_attribute("class", "critical_failed")

        tb = table()
        with tb:
            tb.set_attribute("class", "table_center")
            with tr():
                th("KMP Checks", colspan=6).set_attribute("class", f"detail_rpm")
                th("Kernel Module Checks", colspan=5).set_attribute("class", f"detail_kernel_module")
            with tr():
                th("Name").set_attribute("class", f"detail_0")
                th("Path").set_attribute("class", f"detail_1")
                th("Vendor").set_attribute("class", f"detail_2")
                th(raw("Signature<span class=\"tooltiptext\">Only check there's a signature or not.</span>")).set_attribute("class", f"detail_3 tooltip")
                th(raw("License<span class=\"tooltiptext\">KMP and it's kernel modules should use open source licenses.</span>")).set_attribute("class", f"detail_4 tooltip")
                th(raw("Weak Module Invoked<span class=\"tooltiptext\">Weak Module is necessary to make 3rd party kernel modules installed for one kernel available to KABI-compatible kernels. </span>")).set_attribute("class", f"detail_5 tooltip")
                th(raw("Licenses<span class=\"tooltiptext\">KMP and it's kernel modules should use open source licenses.</span>")).set_attribute("class", f"detail_6 tooltip")
                th(raw("Signatures<span class=\"tooltiptext\">\"supported\" flag: <br/>  \"yes\": Only supported by SUSE<br/>  \"external\": supported by both SUSE and vendor</span>")).set_attribute("class", f"detail_6 tooltip")
                th(raw("Supported Flag<span class=\"tooltiptext\">\"supported\" flag: <br/>  \"yes\": Only supported by SUSE<br/>  \"external\": supported by both SUSE and vendor</span>")).set_attribute("class", f"detail_6 tooltip")
                th(raw("Symbols<span class=\"tooltiptext\">Symbols check is to check whether the symbols in kernel modules matches the symbols in its package.</span>")).set_attribute("class", f"detail_7 tooltip")
                th(raw("Modalias<span class=\"tooltiptext\">Modalias check is to check whether the modalias in kernel modules matches the modalias in its package.</span>")).set_attribute("class", f"detail_8 tooltip")

            for __, row in df.iterrows():
                with tr() as r:
                    if row["level"] == KMPEvaluation.WARNING:
                        r.set_attribute("class", "important_failed_row")
                    elif row["level"] == KMPEvaluation.ERROR:
                        r.set_attribute("class", "critical_failed_row")
                
                    _create_cell(row["name"])
                    _create_cell(row["path"])
                    _create_cell(row["vendor"])
                    _create_cell(row["signature"])
                    _create_cell(row["license"])
                    _create_cell(row["wm2_invoked"])
                    _create_cell(row["km_licenses"])
                    _create_cell(row["km_signatures"])
                    _create_cell(row["supported_flag"])
                    _create_cell(row["symbols"])
                    _create_cell(row["modalias"])
        
        return tb
    
    def to_html(self, df, file):
        pkg_path = os.path.dirname(__file__)
        jinja_tmpl = f"{pkg_path}/../config/templates"
        file_loader = FileSystemLoader(jinja_tmpl)
        env = Environment(loader=file_loader)

        kmp_tmpl = env.get_template("kmp-report.html.jinja")

        kmp_checks = kmp_tmpl.render(version=get_version(), timestamp=generate_timestamp(),
            summary_table=self._summary_to_html(df),
            rpm_details=self._detail_to_html(df),
        )

        with open(file, "w") as f:
            f.write(kmp_checks)
    
    def _create_xlsx_overview(self, ws):
        et = XlsxTemplate()
        et.set_kmp_overview(ws)

    def _summary_to_xlsx(self, wb, df):
        ws = wb.create_sheet("Vendor Summary")
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        
        render = KMPXlsxStyler()
        render.render_summary(ws)

    def _detail_to_xlsx(self, wb, df):
        ws = wb.create_sheet("KMP Detail")
        df_values = df.copy()
        # format value
        df_values = df_values.drop(['level'], axis=1)
        df_values = df_values.applymap(lambda v: v.get("value", ""))
        # fill the values
        df_values = df_values.astype(str)
        for row in dataframe_to_rows(df_values, index=False, header=False):
            ws.append(row)
        
        ws.insert_rows(1, amount=2)
        
        render = KMPXlsxStyler()
        # create header
        def set_header(pairs):
            for loc in pairs:
                ws[loc] = pairs[loc]
                render.set_header(ws[loc])
        
        set_header({
            'A1': 'KMP Checks',
            'G1': 'Kernel Module Checks',
            'A2': 'Name',
            'B2': 'Path',
            'C2': 'Vendor',
            'D2': 'Signature',
            'E2': 'License',
            'F2': 'Weak Module Invoked',
            'G2': 'Licenses',
            'H2': 'Signatures',
            'I2': 'Supported Flag',
            'J2': 'Symbols',
            'K2': 'Modalias'
        })
        ws.merge_cells('A1:F1')
        ws.merge_cells('G1:K1')
        
        pair = {'A' : "name",
        'B' : "path",
        'C' : "vendor",
        'D' : "signature",
        'E' : 'license',
        'F' : "wm2_invoked",
        'G' : "km_licenses",
        'H' : "km_signatures",
        'I' : "supported_flag",
        'J' : "symbols",
        'K' : "modalias",
        }
        
        data_start_row = 3
        row_count = len(df.index) + data_start_row
        for i in range(data_start_row, row_count):
            # TODO: add row level style.
            row_level = df.at[i-data_start_row, 'level']
            for cell in ws[i]:
                v = df.at[i-data_start_row, pair[cell.column_letter]]
                lev = v.get('level', KMPEvaluation.PASS)
                if lev == KMPEvaluation.PASS:
                    render.normal(cell)
                elif lev == KMPEvaluation.WARNING:
                    render.warning(cell)
                elif lev == KMPEvaluation.ERROR:
                    render.error(cell)
        
        render.set_column_width(ws, {'A': 25, 'B': 90, 'C':18, 'D':30, 'E':15, 'F': 10, 'G': 20, 'H': 10, 'I': 30, 'J': 60, 'K': 40})

    def to_xlsx(self, df, file):
        wb = Workbook()
        self._create_xlsx_overview(wb.active)
        
        sum_table = self._summary(df)
        self._summary_to_xlsx(wb, sum_table)
        self._detail_to_xlsx(wb, df)
        
        wb.save(file)
    
    def to_json(self, df, file):
        df.to_json(file, orient="records")

