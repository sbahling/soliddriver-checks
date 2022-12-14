import os
from copy import copy
from openpyxl import load_workbook
from ..config import SDCConf, get_version, generate_timestamp
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

class XlsxStyler:
    def __init__(self):
        self._styler = SDCConf()
        self._header = {}
        self._data_pass = {}
        self._data_warning = {}
        self._data_error = {}

    def set_header(self, cell):
        cell.font   = copy(self._header['font'])
        cell.border = copy(self._header['border'])
        cell.fill   = copy(self._header['fill'])
    
    def normal(self, cell):
        cell.font   = copy(self._data_pass['font'])
        cell.border = copy(self._data_pass['border'])
        cell.fill   = copy(self._data_pass['fill'])
        
    def warning(self, cell):
        cell.font   = copy(self._data_warning['font'])
        cell.border = copy(self._data_warning['border'])
        cell.fill   = copy(self._data_warning['fill'])
    
    def error(self, cell):
        cell.font   = copy(self._data_error['font'])
        cell.border = copy(self._data_error['border'])
        cell.fill   = copy(self._data_error['fill'])
    
    def set_column_width(self, ws, columns):
        for key in columns:
            ws.column_dimensions[key].width = columns[key]


class KMXlsxStyler(XlsxStyler):
    def __init__(self):
        super().__init__()
        
        (self._header['font'], 
         self._header['border'],
         self._header['fill']) = self._styler.get_km_header()
        
        (
            self._data_pass['font'],
            self._data_pass['border'],
            self._data_pass['fill']
        ) = self._styler.get_km_normal()
        
        (
            self._data_warning['font'],
            self._data_warning['border'],
            self._data_warning['fill']
        ) = self._styler.get_km_warning()
        
        (
            self._data_error['font'],
            self._data_error['border'],
            self._data_error['fill']
        ) = self._styler.get_km_error()


class KMPXlsxStyler(XlsxStyler):
    def __init__(self):
        super().__init__()
        (
            self._header['font'], 
            self._header['border'], 
            self._header['fill']
        ) = self._styler.get_kmp_header()
        
        (
            self._data_pass['font'],
            self._data_pass['border'],
            self._data_pass['fill']
        ) = self._styler.get_kmp_normal()
        
        (
            self._data_warning['font'],
            self._data_warning['border'],
            self._data_warning['fill']
        ) = self._styler.get_kmp_warning()
        
        (
            self._data_error['font'],
            self._data_error['border'],
            self._data_error['fill']
        ) = self._styler.get_kmp_error()
        self._row_pass = {}
        (
            self._row_pass['font'],
            self._row_pass['border'],
            self._row_pass['fill']
        ) = self._styler.get_kmp_row_pass()
    
    def render_summary(self, ws):
        # format header
        for cell in ws[1]:
            cell.font = self._header['font']
            cell.border = self._header['border']
            cell.fill = self._header['fill']
        
        # format all the cells
        for row in range(2, ws.max_row + 1):
            for cell in ws[row]:
                cell.font = self._data_pass['font']
                cell.border = self._data_pass['border']
                cell.fill = self._data_pass['fill']
        
        row_pass = Rule(type='expression', dxf=DifferentialStyle(
            font=self._row_pass['font'],
            border=self._row_pass['border'],
            fill=self._row_pass['fill'],
        ))
        row_pass.formula = ['AND($A2 <> "", VALUE(LEFT($C2, FIND(" ",$C2)-1))=0, VALUE(LEFT($D2, FIND(" ", $D2) - 1)) = 0, VALUE(LEFT($E2, FIND(" ", $E2) - 1)) = 0, VALUE(LEFT($F2, FIND(" ", $F2)-1))=0, VALUE(LEFT($G2, FIND(" ", $G2) - 1))=0, VALUE(LEFT($H2, FIND(" ", $H2) - 1))=0, VALUE(LEFT($I2, FIND(" ", $I2) - 1))=0, VALUE(LEFT($J2, FIND(" ", $J2) - 1))=0)']
        ws.conditional_formatting.add(f"A2:J{ws.max_row}", row_pass)
        
        def get_warning_rule():
            return Rule(type='expression', dxf=DifferentialStyle(
                font=self._data_warning['font'],
                border = self._data_warning['border'],
                fill=self._data_warning['fill']
            ))
        warning = get_warning_rule()
        warning.formula = ['$A2 = ""']
        ws.conditional_formatting.add(f"A2:A{ws.max_row}", warning)
        
        warning = get_warning_rule()
        warning.formula = ['VALUE(LEFT($C2, FIND(" ", $C2) - 1)) <> 0']
        ws.conditional_formatting.add(f"C2:C{ws.max_row}", warning)
        
        warning = get_warning_rule()
        warning.formula = ['VALUE(LEFT($D2, FIND(" ", $D2) - 1)) <> 0']
        ws.conditional_formatting.add(f"D2:D{ws.max_row}", warning)
        
        warning = get_warning_rule()
        warning.formula = ['VALUE(LEFT($E2, FIND(" ", $E2) - 1)) <> 0']
        ws.conditional_formatting.add(f"E2:E{ws.max_row}", warning)
        
        warning = get_warning_rule()
        warning.formula = ['VALUE(LEFT($G2, FIND(" ", $G2) - 1)) <> 0']
        ws.conditional_formatting.add(f"G2:G{ws.max_row}", warning)
        
        warning = get_warning_rule()
        warning.formula = ['VALUE(LEFT($H2, FIND(" ", $H2) - 1)) <> 0']
        ws.conditional_formatting.add(f"H2:H{ws.max_row}", warning)
        
        def get_error_rule():
            return Rule(type='expression', dxf=DifferentialStyle(
                font=self._data_error['font'],
                border=self._data_error['border'],
                fill=self._data_error['fill'],
            ))
        
        error = get_error_rule()
        error.formula = ['VALUE(LEFT($F2, FIND(" ", $F2) - 1)) <> 0']
        ws.conditional_formatting.add(f"F2:F{ws.max_row}", error)
        
        error = get_error_rule()
        error.formula = ['VALUE(LEFT($I2, FIND(" ", $I2) - 1)) <> 0']
        ws.conditional_formatting.add(f"I2:I{ws.max_row}", error)
        
        error = get_error_rule()
        error.formula = ['VALUE(LEFT($J2, FIND(" ", $J2) - 1)) <> 0']
        ws.conditional_formatting.add(f"J2:J{ws.max_row}", error)
        
        self.set_column_width(ws, {
            'A': 40,
            'B': 15,
            'C': 15,
            'D': 15,
            'E': 15,
            'F': 15,
            'G': 15,
            'H': 15,
            'I': 15,
            'J': 15
        })

class XlsxTemplate:
    def __init__(self):
        pkg_path = os.path.dirname(__file__)
        self._cfg_path = f"{pkg_path}/../config/templates/templates.xlsx"

    def set_km_overview(self, ws):
        self._copy_work_sheep("km-report-overview", ws)

    def set_kmp_overview(self, ws):
        self._copy_work_sheep("kmp-report-overview", ws)

    def _copy_work_sheep(self, title, ws):
        tmpl = load_workbook(filename=self._cfg_path)
        cover = tmpl[title]
        ws.title = "Overview"
        
        for row in cover.rows:
            for cell in row:
                ws[cell.coordinate].value     = copy(cell.value)
                ws[cell.coordinate].style     = copy(cell.style)
                ws[cell.coordinate].font      = copy(cell.font)
                ws[cell.coordinate].border    = copy(cell.border)
                ws[cell.coordinate].fill      = copy(cell.fill)
                ws[cell.coordinate].alignment = copy(cell.alignment)

        ws["A5"].value = get_version()
        ws["A6"].value = generate_timestamp()
        ws.column_dimensions['A'].width = 200
        ws.page_setup.fitToHeight = 1