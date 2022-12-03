from sqlite3 import Timestamp
import pandas as pd
import os
from pathlib import Path
from dominate.tags import html, body, h1, div, tr, td, th, table, style, ul, li, p
from dominate.util import raw
from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment,
    NamedStyle,
    borders,
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import string
from openpyxl import Workbook
from jinja2 import Environment, FileSystemLoader
import re
from copy import copy
from ..config import SDCConf, ExcelTemplate, get_version, generate_timestamp
from .data_analysis import KMPEvaluation

class DriversExporter:
    def __init__(self):
        self._style = SDCConf()

    def _driver_path_check(self, driver_path):
        # if the driver is located in weak-updates directory,
        # check wether it's valid or not first.
        for i, wud_checks in self._wu_dt.iterrows():
            if driver_path == wud_checks["driver"] and wud_checks["result"] == "Failed":
                return None

        # TODO: this should be optimized!
        result = re.match(
            r"^/lib/modules/[0-9]+.[0-9]+.[0-9]+\-[0-9]+\-[a-z]+/(updates/|weak-updates/|extra/)",
            driver_path,
        )

        if result is not None:
            return result
        else:
            return re.match(
                r"^/lib/modules/[0-9]+.[0-9]+.[0-9]+\-[0-9]+.[0-9]+\-[a-z]+/(updates/|weak-updates/|extra/)",
                driver_path,
            )

    def _format_row_html(self, row):
        path = row["path"]
        supported = row["flag_supported"]
        supported = supported.split(" ")
        license = row["license"]
        signature = row["signature"]
        rpm = row["rpm"]
        vld_lic = self._style.get_valid_licenses()

        critical_style = self._style.get_driver_html_warn_critical()
        important_style = self._style.get_driver_html_warn_important()
        cs_bgColor = critical_style["background-color"]
        cs_color = critical_style["color"]
        cs_border = critical_style["border"]

        is_bgColor = important_style["background-color"]
        is_border = important_style["border"]

        warn_level = 0  # 0: normal, 1: important, 2: critical

        name_style = ""
        path_style = ""
        supported_style = ""
        license_style = ""
        signature_style = ""
        suse_release_style = ""
        running_style = ""
        rpm_style = ""

        if not ValidLicense(license, vld_lic):
            warn_level = 1
            license_style = f"background-color:{is_bgColor}"

        if not signature:
            warn_level = 1
            signature_style = f"background-color:{is_bgColor}"

        if not self._driver_path_check(path):
            warn_level = 2
            path_style = f"background-color:{cs_bgColor} color:{cs_color}"

        if (len(supported) == 0) or (
            len(supported) == 1 and supported[0] != "external"
        ):
            warn_level = 2
            supported_style = f"background-color:{cs_bgColor} color:{cs_color}"
        if len(supported) > 1:
            v1 = supported[0]
            if v1 != "external":
                warn_level = 2
                supported_style = f"background-color:{cs_bgColor} color:{cs_color}"
            else:
                warn_level = 1
                supported_style = f"background-color:{is_bgColor}"
                for v in supported:
                    if v != v1 and v != "external":
                        warn_level = 2
                        supported_style = (
                            f"background-color:{cs_bgColor} color:{cs_color}"
                        )

        if "is not owned by any package" in rpm:
            warn_level = 2
            rpm_style = f"background-color:{cs_bgColor} color:{cs_color}"

        row_style = [
            name_style,
            path_style,
            supported_style,
            license_style,
            signature_style,
            suse_release_style,
            running_style,
            rpm_style,
        ]
        return row_style

    def to_json(self, driver_tables, file):
        jf = dict()
        for label, driver_table in driver_tables.items():
            if driver_table is not None:
                drivers = driver_table["drivers"]
                wu_drivers = driver_table["weak-update-drivers"]
                drivers_buff = drivers.to_json(orient="records")
                wu_drviers_buff = wu_drivers.to_json(orient="records")
                noinfo_drivers = json.dumps(driver_table["noinfo-drivers"])
            else:
                drivers_buff = "{}"
                wu_drviers_buff = "{}"

            jf[label] = {
                "drivers": json.loads(drivers_buff),
                "weak-update-drivers": json.loads(wu_drviers_buff),
                "noinfo-drivers": json.loads(noinfo_drivers),
            }

        with open(file, "w") as fp:
            json.dump(jf, fp)

    def _get_third_party_drivers(self, drivers):
        df = drivers.copy()
        sig_keys = self._style.get_driver_sig_keys()
        keys = []
        for k in sig_keys:
            keys.append(k["key"])

        ser = ~df.rpm_sig_key.isin(keys)
        df = df[ser]
        df.drop("rpm_sig_key", axis=1, inplace=True)

        return df

    def _refmt_supported(self, drivers):
        df = drivers.copy()
        for i, row in df.iterrows():
            row["flag_supported"] = " ".join(row["flag_supported"])

        return df

    def _append_noinfo_drivers(self, noinfo_drivers, drivers):
        for driver in noinfo_drivers:
            row = [
                driver,
                "Can not find file under /lib/modules",
                "",
                "",
                "",
                "",
                "True",
                f"driver {driver} is not owned by any package",
            ]
            drivers = drivers.append(
                pd.Series(
                    row,
                    index=[
                        "name",
                        "path",
                        "flag_supported",
                        "license",
                        "signature",
                        "os-release",
                        "running",
                        "rpm",
                    ],
                ),
                ignore_index=True,
            )

        return drivers

    def to_html(self, driver_tables, file):
        pkg_path = os.path.dirname(__file__)
        jinja_tmpl = f"{pkg_path}/../config/templates"
        file_loader = FileSystemLoader(jinja_tmpl)
        env = Environment(loader=file_loader)

        driver_tmpl = env.get_template("driver-checks.html.jinja")

        details = []
        for label, driver_table in driver_tables.items():
            if driver_table is None:
                details.append({"name": label, "table": "Connect error!"})
                continue

            dt = driver_table["drivers"]
            self._wu_dt = driver_table["weak-update-drivers"]
            total_drivers, tp_drivers, failed_drivers = self._get_server_summary(dt)
            df = dt.copy()
            df = self._get_third_party_drivers(df)

            # add no information drivers
            noinfo_drivers = driver_table["noinfo-drivers"]
            df = self._append_noinfo_drivers(noinfo_drivers, df)
            df.loc[df["running"] == "True", "running"] = "&#9989;"
            df.loc[df["running"] == "False", "running"] = "&#9940;"
            df = self._refmt_supported(df)
            ts = (
                # df.style.hide(axis='index')
                df.style.hide_index()
                .set_table_attributes('class="table_center"')
                .apply(self._format_row_html, axis=1)
            )

            details.append(
                {
                    "name": label,
                    "total_drivers": total_drivers,
                    "third_party_drivers": tp_drivers,
                    "failed_drivers": failed_drivers,
                    "table": ts.render(),
                }
            )

        driver_checks = driver_tmpl.render(version=_get_version(), timestamp=_generate_timestamp(), details=details)

        with open(file, "w") as f:
            f.write(driver_checks)

    def _xlsx_create_overview(self, wb):
        et = ExcelTemplate()
        et.set_driver_check_overview(wb.active)

    def _get_failed_driver_count(self, df):
        count = 0
        vld_lic = self._style.get_valid_licenses()
        for i, row in df.iterrows():
            if (
                "is not owned by any package" in row["rpm"]
                or " ".join(row["flag_supported"]) != "external"
                or str(row["signature"]) != "True"
                or not self._driver_path_check(row["path"])
                or not ValidLicense(row["license"], vld_lic)
            ):
                count += 1

        return count

    def _get_server_summary(self, driver_table):
        total_drivers = len(driver_table.index)
        third_party_drivers = self._get_third_party_drivers(driver_table)
        tpd_count = len(third_party_drivers.index)

        failed_count = self._get_failed_driver_count(third_party_drivers)

        return total_drivers, tpd_count, failed_count

    def _xlsx_create_table(self, wb, label, driver_tables):
        ws_dc = wb.create_sheet(label)
        if driver_tables is None:
            return

        dt = driver_tables["drivers"]
        self._wu_dt = driver_tables["weak-update-drivers"]
        noinfo_drivers = driver_tables["noinfo-drivers"]
        df = self._get_third_party_drivers(dt)
        noinfo_drivers = driver_tables["noinfo-drivers"]
        df = self._append_noinfo_drivers(noinfo_drivers, df)
        df = self._refmt_supported(df)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws_dc.append(row)

        (
            header_font,
            header_border,
            header_fill,
        ) = self._style.get_driver_xslx_table_header()
        for cell in ws_dc[1]:
            cell.font = header_font
            cell.border = header_border
            cell.fill = header_fill

        (
            ctc_font,
            ctc_border,
            ctc_fill,
        ) = self._style.get_driver_xslx_table_critical_failed()
        (
            imt_font,
            imt_border,
            imt_fill,
        ) = self._style.get_driver_xslx_table_important_failed()
        (
            normal_font,
            normal_border,
            normal_fill,
        ) = self._style.get_driver_xslx_table_normal()
        vld_lic = self._style.get_valid_licenses()

        last_record_row_no = len(df.index) + 2  # the title of the table also count
        for i in range(2, last_record_row_no):
            # default format.
            for cell in ws_dc[i]:
                cell.font = normal_font
                cell.border = normal_border
                cell.fill = normal_fill

            path = ws_dc[f"B{i}"].value
            flag = ws_dc[f"C{i}"].value.split(" ")
            license = ws_dc[f"D{i}"].value
            sig = ws_dc[f"E{i}"].value
            rpm = ws_dc[f"H{i}"].value

            w_level = 0

            if not ValidLicense(license, vld_lic):
                w_level = 1
                ws_dc[f"D{i}"].font = imt_font
                ws_dc[f"D{i}"].border = imt_border
                ws_dc[f"D{i}"].fill = imt_fill

            if sig == "":
                w_level = 1
                ws_dc[f"E{i}"].font = imt_font
                ws_dc[f"E{i}"].border = imt_border
                ws_dc[f"E{i}"].fill = imt_fill

            if not self._driver_path_check(path):
                w_level = 2
                ws_dc[f"B{i}"].font = imt_font
                ws_dc[f"B{i}"].border = imt_border
                ws_dc[f"B{i}"].fill = imt_fill

            if (len(flag) == 0) or (len(flag) == 1 and flag[0] != "external"):
                w_level = 2
                ws_dc[f"C{i}"].font = ctc_font
                ws_dc[f"C{i}"].border = ctc_border
                ws_dc[f"C{i}"].fill = ctc_fill
            elif len(flag) > 1:
                v1 = flag[0]
                if v1 != "external":
                    warn_level = 2
                    ws_dc[f"C{i}"].font = ctc_font
                    ws_dc[f"C{i}"].border = ctc_border
                    ws_dc[f"C{i}"].fill = ctc_fill
                else:
                    warn_level = 1
                    ws_dc[f"C{i}"].font = imt_font
                    ws_dc[f"C{i}"].border = imt_border
                    ws_dc[f"C{i}"].fill = imt_fill
                    for v in flag:
                        if v != v1:
                            warn_level = 2
                            ws_dc[f"C{i}"].font = ctc_font
                            ws_dc[f"C{i}"].border = ctc_border
                            ws_dc[f"C{i}"].fill = ctc_fill
                            break

            if "is not owned by any package" in rpm:
                warn_level = 2
                ws_dc[f"H{i}"].font = ctc_font
                ws_dc[f"H{i}"].border = ctc_border
                ws_dc[f"H{i}"].fill = ctc_fill

    def to_excel(self, driver_tables, file):
        wb = Workbook()
        self._xlsx_create_overview(wb)

        for label, dt in driver_tables.items():
            self._xlsx_create_table(wb, label, dt)

        if os.path.exists(file):
            os.remove(file)

        wb.save(file)

    def to_pdf(self, driver_tables, file):
        self.to_html(driver_tables, ".tmp.rpms.html")

        content = ""
        with open(".tmp.rpms.html", "r") as fp:
            content = fp.read()

        content = str(content).replace("&#9989;", "True").replace("&#9940;", "False")
        with open(".tmp.rpms.html", "w") as fp:
            fp.write(content)

        pdfkit.from_file(".tmp.rpms.html", file)
        os.remove(".tmp.rpms.html")

    def to_all(self, driver_tables, directory):
        excel_file = os.path.join(directory, "check_result.xlsx")
        html_file = os.path.join(directory, "check_result.html")
        pdf_file = os.path.join(directory, "check_result.pdf")
        json_file = os.path.join(directory, "check_result.json")

        if not os.path.exists(directory):
            os.mkdir(directory)
        else:
            if os.path.exists(excel_file):
                os.remove(excel_file)
            if os.path.exists(html_file):
                os.remove(html_file)
            if os.path.exists(pdf_file):
                os.remove(pdf_file)
            if os.path.exists(json_file):
                os.remove(json_file)

        self.to_excel(driver_tables, excel_file)
        self.to_html(driver_tables, html_file)
        self.to_pdf(driver_tables, pdf_file)
        self.to_pdf(driver_tables, json_file)
