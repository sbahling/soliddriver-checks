import os
from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
)
from openpyxl import load_workbook
import json
from .version import __VERSION__
from datetime import datetime
from copy import copy


def get_version():
    return f"version: {__VERSION__}"


def generate_timestamp():
    return f"timestamp: {datetime.now()}"


class SDCConf:
    def __init__(self):
        pkg_path = os.path.dirname(__file__)
        cfg_path = f"{pkg_path}/config/soliddriver-checks.conf"

        with open(cfg_path, "r") as fp:
            self._conf = json.load(fp)

    def _get_xlsx_info(self, *locs):
        conf = self._conf
        for loc in locs:
            conf = conf[loc]

        font = Font(
            name=conf["font"]["family"],
            size=conf["font"]["size"],
            bold=conf["font"]["bold"],
            color=conf["font"]["color"],
        )
        sd = Side(
            border_style=conf["side"]["border_style"],
            color=conf["side"]["color"]
        )
        bd = Border(top=sd, left=sd, right=sd, bottom=sd)
        fill = PatternFill(
            start_color=conf["fill"]["bgcolor"],
            end_color=conf["fill"]["bgcolor"],
            fill_type="solid",
        )

        return font, bd, fill

    def get_valid_licenses(self):
        return self._conf["valid-licenses"]

    def get_kmp_header(self):
        return self._get_xlsx_info("kmp", "xlsx", "table", "header")

    def get_kmp_normal(self):
        return self._get_xlsx_info("kmp", "xlsx", "table", "data", "normal")

    def get_kmp_warning(self):
        return self._get_xlsx_info(
            "kmp", "xlsx", "table", "data", "warning"
        )

    def get_kmp_error(self):
        return self._get_xlsx_info(
            "kmp", "xlsx", "table", "data", "error"
        )

    def get_kmp_row_pass(self):
        return self._get_xlsx_info("kmp", "xlsx", "table", "row", "pass")

    def get_km_sig_keys(self):
        return self._conf["km"]["sig-keys"]

    def get_km_header(self):
        return self._get_xlsx_info("km", "xlsx", "table", "header")

    def get_km_normal(self):
        return self._get_xlsx_info("km", "xlsx", "table", "data", "normal")

    def get_km_warning(self):
        return self._get_xlsx_info(
            "km", "xlsx", "table", "data", "warning"
        )

    def get_km_error(self):
        return self._get_xlsx_info(
            "km", "xlsx", "table", "data", "error"
        )

    def get_km_warn_row(self):
        return self._get_xlsx_info("km", "xlsx", "table", "row", "warn")

    def get_km_html_warning(self):
        return self._conf["km"]["html"]["warning"]

    def get_km_html_error(self):
        return self._conf["km"]["html"]["error"]


class xlsxTemplate:
    def __init__(self):
        pkg_path = os.path.dirname(__file__)
        self._cfg_path = f"{pkg_path}/../config/templates/templates.xlsx"

    def set_km_check_overview(self, ws):
        self._copy_work_sheep("km-report-overview", ws)

    def set_kmp_check_overview(self, ws):
        self._copy_work_sheep("kmp-report-overview", ws)

    def _copy_work_sheep(self, title, ws):
        tmpl = load_workbook(filename=self._cfg_path)
        cover = tmpl[title]

        ws.title = "Overview"
        for row in cover.rows:
            for cell in row:
                ws[cell.coordinate].value = copy(cell.value)
                # ws[cell.coordinate].style = copy(cell.style)
                ws[cell.coordinate].font = copy(cell.font)
                ws[cell.coordinate].border = copy(cell.border)
                ws[cell.coordinate].fill = copy(cell.fill)
                ws[cell.coordinate].alignment = copy(cell.alignment)

        ws["A5"].value = get_version()
        ws["A6"].value = generate_timestamp()
        ws.column_dimensions['A'].width = 200
